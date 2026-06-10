# -*- coding: utf-8 -*-
"""
报货 PDF 自动汇总 GUI 工具

功能：
1. 读取 Temu/拣货单 PDF 中的 属性集、SKU ID、SKU货号、实际发货数。
2. 按自定义规则把 SKU货号 映射为你表格里的货号。
3. 支持套装拆分：一个 SKU 可拆成多个款号。
4. 支持同款多颜色拆分：白色+卡其色-L 可拆成 白色-L、卡其色-L。
5. 自动合并统计：输出款号 + 颜色尺码 相同则数量相加。
6. GUI 可维护规则，并保存到 JSON 配置。

依赖：
    pip install pymupdf openpyxl

GUI 启动：
    python baohuo_pdf_gui.py

命令行批处理：
    python baohuo_pdf_gui.py --cli TEST.pdf -o TEST_报货汇总.xlsx
    python baohuo_pdf_gui.py --cli ./pdf_folder -o ./输出目录
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import traceback
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import fitz  # PyMuPDF
except ImportError as e:
    raise SystemExit("缺少 PyMuPDF，请先运行：pip install pymupdf") from e

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError as e:
    raise SystemExit("缺少 openpyxl，请先运行：pip install openpyxl") from e

# Tkinter 在部分精简 Python 环境中可能不存在；只有 GUI 模式才真正需要。
try:
    import tkinter as tk
    import tkinter.font as tkfont
    from tkinter import ttk, filedialog, messagebox
except Exception:  # pragma: no cover
    tk = None
    tkfont = None
    ttk = None
    filedialog = None
    messagebox = None

APP_NAME = "PDF_Checker"
CONFIG_FILE_NAME = "check_rules.json"
APPDATA_DIR_NAME = "PDF_Checker"
LEGACY_APPDATA_DIR_NAMES = ["BaohuoPDFSummaryTool"]

HAN_RE = re.compile(r"[\u4e00-\u9fff]")
# 可识别：SK-010、PG-JS-836、MGTZ-26、MG-55、SSL-14、XZ-21+XZ-20、MG-55+SSL-14
STYLE_CODE_RE = re.compile(r"^([A-Z]+(?:-[A-Z]+)*-\d+(?:\+[A-Z]+(?:-[A-Z]+)*-\d+)*)", re.I)
NUMERIC_STYLE_CODE_RE = re.compile(r"^(\d+)(?:-[A-Z][A-Z0-9]*(?:\+[A-Z][A-Z0-9]*)*)?$", re.I)
SIZE_ORDER = ["XS", "S", "M", "L", "XL", "XXL", "XXXL", "XXXXL"]
COLOR_ORDER = [
    "白色", "黑色", "灰色", "浅灰色", "深灰色",
    "浅蓝色", "天蓝色", "藏青色", "军绿色", "军绿", "卡其色", "杏色", "粉红色", "土黄色",
    "白色+卡其色", "白色+黑色", "黑色+藏青色", "军绿+卡其色", "军绿色+卡其色",
]

COLOR_ALIASES = {
    "浅灰": "浅灰色",
    "深灰": "深灰色",
    "浅蓝": "浅蓝色",
    "天蓝": "天蓝色",
    "藏青": "藏青色",
    "军绿": "军绿色",
    "卡其": "卡其色",
    "粉红": "粉红色",
    "土黄": "土黄色",
}

DEFAULT_CONFIG: Dict[str, Any] = {
    "version": 3,
    "general": {
        "style_prefix_replace": [
            {"from": "MGTZ-", "to": "MG-", "enabled": True},
        ],
        "style_code_replace": [],
        "quantity_only_styles": [],
        "quantity_only_group_name": "数量统计",
        "split_auto_combo_styles": True,
        "merge_same_style_attr": True,
        "default_output_name": "报货汇总.xlsx",
        "default_group_name": "未分组",
        "show_group_headers": True,
        "blank_row_between_styles": True,
    },
    "color_map": {
        "White": "白色",
        "Black": "黑色",
        "Grey": "灰色",
        "Gray": "灰色",
        "LightGrey": "浅灰色",
        "LightGray": "浅灰色",
        "DarkGrey": "深灰色",
        "DarkGray": "深灰色",
        "LightBlue": "浅蓝色",
        "SkyBlue": "天蓝色",
        "NavyBlue": "藏青色",
        "DarkBlue": "藏青色",
        "ArmyGreen": "军绿色",
        "ArmyBlue": "军绿色",
        "Khaki": "卡其色",
        "Apricot": "杏色",
        "Red": "红色",
        "PinkRed": "粉红色",
        "EarthyYellow": "土黄色",
    },
    "rules": [
        {
            "enabled": True,
            "name": "默认：自动识别款号，组合款自动拆款",
            "priority": 1000,
            "match_type": "regex",
            "pattern": ".*",
            "output_styles": "{auto}",
            "color_mode": "keep",  # keep / split_attr_color / fixed_colors / sku_color
            "fixed_colors": "",
            "qty_multiplier": 1,
            "group_name": "自动识别",
            "note": "兜底规则。例：MG-55+SSL-14-Black 自动拆成 MG-55 与 SSL-14。",
        },
        {
            "enabled": False,
            "name": "示例：同款双色拆分 MG-55-White+Khaki",
            "priority": 10,
            "match_type": "contains",
            "pattern": "MG-55-White+Khaki",
            "output_styles": "MG-55",
            "color_mode": "sku_color",
            "fixed_colors": "",
            "qty_multiplier": 1,
            "group_name": "MG系列",
            "note": "启用后会把 MG-55-White+Khaki 拆成 MG-55 白色-尺码 与 MG-55 卡其色-尺码。",
        },
        {
            "enabled": False,
            "name": "示例：固定拆两个款式",
            "priority": 20,
            "match_type": "contains",
            "pattern": "MG-55+SSL-14-Black",
            "output_styles": "MG-55,SSL-14",
            "color_mode": "keep",
            "fixed_colors": "",
            "qty_multiplier": 1,
            "group_name": "套装拆分",
            "note": "启用后会把这个 SKU 同时计入 MG-55 与 SSL-14。",
        },
        {
            "enabled": False,
            "name": "示例：按属性集拆色",
            "priority": 30,
            "match_type": "contains",
            "pattern": "White+Black",
            "output_styles": "{auto}",
            "color_mode": "split_attr_color",
            "fixed_colors": "",
            "qty_multiplier": 1,
            "group_name": "双色拆分",
            "note": "例：白色+黑色-XL 拆成 白色-XL 与 黑色-XL。",
        },
    ],
    "group_rules": [
        {
            "enabled": True,
            "priority": 100,
            "name": "默认：MG系列",
            "match_type": "startswith",
            "pattern": "MG-",
            "group_name": "MG系列",
        },
        {
            "enabled": True,
            "priority": 110,
            "name": "默认：SSL系列",
            "match_type": "startswith",
            "pattern": "SSL-",
            "group_name": "SSL系列",
        },
        {
            "enabled": True,
            "priority": 120,
            "name": "默认：XZ系列",
            "match_type": "startswith",
            "pattern": "XZ-",
            "group_name": "XZ系列",
        },
        {
            "enabled": True,
            "priority": 130,
            "name": "默认：SK系列",
            "match_type": "startswith",
            "pattern": "SK-",
            "group_name": "SK系列",
        },
        {
            "enabled": True,
            "priority": 140,
            "name": "默认：PG系列",
            "match_type": "startswith",
            "pattern": "PG-",
            "group_name": "PG系列",
        },
    ],
}


@dataclass
class ParsedRow:
    source_file: str
    page: int
    attr: str
    sku_id: str
    sku_code: str
    qty: int


@dataclass
class OutputRow:
    source_file: str
    page: int
    original_attr: str
    sku_id: str
    sku_code: str
    output_style: str
    group_name: str
    output_attr: str
    qty: float
    rule_name: str
    dedupe_existing_attrs: bool = False


@dataclass
class ParseIssue:
    source_file: str
    page: Optional[int]
    issue: str
    text: str


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def default_config_path() -> Path:
    return user_config_dir() / CONFIG_FILE_NAME


def bundled_config_path() -> Path:
    bundle_dir = getattr(sys, "_MEIPASS", None)
    if bundle_dir:
        bundled = Path(bundle_dir) / CONFIG_FILE_NAME
        if bundled.exists():
            return bundled
    return app_dir() / CONFIG_FILE_NAME


def user_config_dir() -> Path:
    if sys.platform == "win32":
        root = os.environ.get("APPDATA") or os.environ.get("LOCALAPPDATA")
        if root:
            return Path(root) / APPDATA_DIR_NAME
    return Path.home() / f".{APPDATA_DIR_NAME}"


def legacy_config_paths() -> List[Path]:
    paths: List[Path] = []
    if sys.platform == "win32":
        root = os.environ.get("APPDATA") or os.environ.get("LOCALAPPDATA")
        if root:
            paths.extend(Path(root) / name / CONFIG_FILE_NAME for name in LEGACY_APPDATA_DIR_NAMES)
    else:
        paths.extend(Path.home() / f".{name}" / CONFIG_FILE_NAME for name in LEGACY_APPDATA_DIR_NAMES)
    return paths


def deep_copy_default_config() -> Dict[str, Any]:
    return json.loads(json.dumps(DEFAULT_CONFIG, ensure_ascii=False))


def load_config(path: Optional[Path] = None) -> Dict[str, Any]:
    path = path or default_config_path()
    if not path.exists():
        source_path = next((p for p in legacy_config_paths() if p.exists()), None)
        template_path = bundled_config_path()
        if source_path is None and path != template_path and template_path.exists():
            source_path = template_path
        if source_path is not None:
            with open(source_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        else:
            cfg = deep_copy_default_config()
        save_config(cfg, path)
        return cfg
    with open(path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    # 兼容旧配置缺字段
    merged = deep_copy_default_config()
    merged.update(cfg)
    merged["general"].update(cfg.get("general", {}))
    merged["color_map"].update(cfg.get("color_map", {}))
    merged["rules"] = cfg.get("rules", merged["rules"])
    merged["group_rules"] = cfg.get("group_rules", merged.get("group_rules", []))
    return merged


def save_config(cfg: Dict[str, Any], path: Optional[Path] = None) -> None:
    path = path or default_config_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def clean_text_line(s: str) -> str:
    return (s or "").strip().replace("\u3000", " ")


def is_attr_line(s: str) -> bool:
    if not s or "-" not in s:
        return False
    if not HAN_RE.search(s):
        return False
    bad_prefix = ("SKC货号", "备货", "要求", "创建", "收货", "打印", "第", "数量", "合计", "序号", "商品信息")
    return not s.startswith(bad_prefix)


def is_sku_id(s: str) -> bool:
    return bool(re.fullmatch(r"\d{8,14}", s or ""))


def is_int(s: str) -> bool:
    return bool(re.fullmatch(r"\d+", s or ""))


def detect_report_date(lines_with_page: List[Tuple[int, str]]) -> str:
    full_text = "\n".join(line for _, line in lines_with_page)
    m = re.search(r"创建时间：(\d{4})-(\d{2})-(\d{2})", full_text)
    if not m:
        m = re.search(r"要求发货时间：(\d{4})-(\d{2})-(\d{2})", full_text)
    if not m:
        return "报货汇总"
    return f"{int(m.group(2))}月{int(m.group(3))}号"


def extract_pdf_lines(pdf_path: Path) -> List[Tuple[int, str]]:
    lines: List[Tuple[int, str]] = []
    with fitz.open(pdf_path) as doc:
        for page_no, page in enumerate(doc, start=1):
            text = page.get_text("text")
            for line in text.splitlines():
                line = clean_text_line(line)
                if line:
                    lines.append((page_no, line))
    return lines


def parse_pdf_rows(pdf_path: Path) -> Tuple[List[ParsedRow], List[ParseIssue], str]:
    lines_with_page = extract_pdf_lines(pdf_path)
    lines = [line for _, line in lines_with_page]
    pages = [page for page, _ in lines_with_page]
    rows: List[ParsedRow] = []
    issues: List[ParseIssue] = []

    i = 0
    while i < len(lines):
        line = lines[i]
        if is_attr_line(line):
            attr = line
            j = i + 1

            # 处理属性集换行，例如 白色+卡其色-XX / XL -> 白色+卡其色-XXXL
            while j < len(lines) and not is_sku_id(lines[j]):
                if re.fullmatch(r"[XSML]+", lines[j]):
                    attr += lines[j]
                    j += 1
                else:
                    break

            if j < len(lines) and is_sku_id(lines[j]):
                sku_id = lines[j]
                k = j + 1
                sku_parts: List[str] = []

                # 常规格式：SKU货号是字母/符号款号，下一行纯数字是数量。
                # 兼容纯数字款号：SKU ID 后若出现连续两个纯数字，第一行是款号，第二行是数量。
                if k + 1 < len(lines) and is_int(lines[k]) and is_int(lines[k + 1]):
                    sku_code = lines[k].strip()
                    qty = int(lines[k + 1])
                    qty_line_index = k + 1
                else:
                    while k < len(lines) and not is_int(lines[k]):
                        sku_parts.append(lines[k])
                        k += 1
                    if not (k < len(lines) and sku_parts and is_int(lines[k])):
                        issues.append(ParseIssue(pdf_path.name, pages[i], "疑似属性行但未读取到 SKU货号/数量", line))
                        i += 1
                        continue
                    sku_code = "".join(sku_parts).strip()
                    qty = int(lines[k])
                    qty_line_index = k

                if sku_code:
                    rows.append(ParsedRow(
                        source_file=pdf_path.name,
                        page=pages[i],
                        attr=attr,
                        sku_id=sku_id,
                        sku_code=sku_code,
                        qty=qty,
                    ))
                    i = qty_line_index + 1
                    continue
        i += 1

    title = detect_report_date(lines_with_page)
    if not rows:
        issues.append(ParseIssue(pdf_path.name, None, "未解析到数据", "请确认 PDF 不是纯图片扫描件，且包含 属性集 / SKU ID / SKU货号 / 实际发货数字段。"))
    return rows, issues, title


def normalize_style(style: str, cfg: Dict[str, Any]) -> str:
    style = (style or "").strip()
    if re.fullmatch(r"[A-Z]+(?:-[A-Z]+)*-\d+(?:\+[A-Z]+(?:-[A-Z]+)*-\d+)*", style, flags=re.I):
        style = style.upper()
    for item in cfg.get("general", {}).get("style_prefix_replace", []):
        if not item.get("enabled", True):
            continue
        old = str(item.get("from", ""))
        new = str(item.get("to", ""))
        if old and style.startswith(old):
            style = new + style[len(old):]
    for item in cfg.get("general", {}).get("style_code_replace", []):
        if not item.get("enabled", True):
            continue
        old = str(item.get("from", "")).strip()
        new = str(item.get("to", "")).strip()
        if old and new and style == old:
            style = new
            break
    return style


def quantity_only_group_name(cfg: Dict[str, Any]) -> str:
    name = str(cfg.get("general", {}).get("quantity_only_group_name", "数量统计") or "数量统计").strip()
    return name or "数量统计"


def quantity_only_style_set(cfg: Dict[str, Any]) -> set:
    styles = cfg.get("general", {}).get("quantity_only_styles", [])
    result = set()
    if isinstance(styles, str):
        styles = normalize_separator_list(styles)
    for style in styles:
        normalized = normalize_style(str(style), cfg).strip()
        if normalized:
            result.add(normalized)
    return result


def is_quantity_only_style(style: str, cfg: Dict[str, Any]) -> bool:
    return normalize_style(style, cfg) in quantity_only_style_set(cfg)


def is_attrless_quantity_style(style: str, cfg: Dict[str, Any]) -> bool:
    normalized = normalize_style(style, cfg)
    return re.fullmatch(r"\d+", normalized) is not None and normalized in quantity_only_style_set(cfg)


def should_dedupe_existing_attrs(rule: Dict[str, Any]) -> bool:
    if rule.get("dedupe_existing_attrs", False):
        return True
    pattern = str(rule.get("pattern", "")).upper()
    output_styles = str(rule.get("output_styles", "")).upper()
    return rule.get("color_mode") == "sku_color" and ("MG-8" in pattern or "MG-8" in output_styles)


def should_split_mg8_combo(row: ParsedRow, styles: List[str], cfg: Dict[str, Any]) -> bool:
    if not any(normalize_style(style, cfg) == "MG-8" for style in styles):
        return False
    color, _ = split_attr(row.attr)
    return "+" in color


def get_auto_style_codes(sku_code: str, cfg: Dict[str, Any]) -> List[str]:
    code = (sku_code or "").strip().replace(" ", "-")
    code = re.sub(r"-+", "-", code)
    m = STYLE_CODE_RE.match(code)
    if not m:
        numeric_m = NUMERIC_STYLE_CODE_RE.match(code)
        if numeric_m:
            return [normalize_style(numeric_m.group(1), cfg)]
        return [normalize_style(code, cfg)] if code else ["未识别款号"]
    base = m.group(1)
    split_combo = cfg.get("general", {}).get("split_auto_combo_styles", True)
    if split_combo and "+" in base:
        return [normalize_style(x, cfg) for x in base.split("+") if x]
    return [normalize_style(base, cfg)]


def split_attr(attr: str) -> Tuple[str, str]:
    attr = (attr or "").strip()
    if "-" not in attr:
        return attr, ""
    color, size = attr.rsplit("-", 1)
    return color, size


def normalize_color_name(color: str) -> str:
    color = (color or "").strip()
    if not color:
        return color
    parts = [part.strip() for part in re.split(r"\+", color)]
    normalized = [COLOR_ALIASES.get(part, part) for part in parts if part]
    return "+".join(normalized) if normalized else color


def normalize_output_attr(attr: str) -> str:
    color, size = split_attr(attr)
    color = normalize_color_name(color)
    if size:
        return f"{color}-{size}" if color else size
    return color


def normalize_separator_list(s: str) -> List[str]:
    if not s:
        return []
    s = s.replace("，", ",").replace("；", ",").replace(";", ",").replace("/", ",")
    parts = []
    for chunk in s.split(","):
        chunk = chunk.strip()
        if chunk:
            parts.append(chunk)
    return parts


def parse_output_styles(rule_styles: str, sku_code: str, cfg: Dict[str, Any]) -> List[str]:
    rule_styles = (rule_styles or "").strip()
    if not rule_styles or rule_styles == "{auto}":
        return get_auto_style_codes(sku_code, cfg)
    # 用逗号分隔。允许用户写 MG-55,SSL-14；不要用 + 分隔，避免和款号中的 + 混淆。
    styles = normalize_separator_list(rule_styles)
    if not styles:
        return get_auto_style_codes(sku_code, cfg)
    return [normalize_style(x, cfg) for x in styles]


def english_color_tokens_from_sku(sku_code: str) -> List[str]:
    """从 SKU 末尾颜色段里取英文颜色。例：MG-55-White+Khaki -> White, Khaki。"""
    code = (sku_code or "").strip().replace(" ", "-")
    m = STYLE_CODE_RE.match(code)
    if not m:
        return []
    rest = code[m.end():]
    rest = rest.strip("-")
    if not rest:
        return []
    # 如果 SKU 末尾还有尺码，例如 JB-003-Black-XXXL，把末尾尺码去掉。
    rest = re.sub(r"-(XS|S|M|L|XL|XXL|XXXL|XXXXL)$", "", rest, flags=re.I)
    tokens = []
    for p in re.split(r"\+|,|/", rest):
        p = p.strip("-").strip()
        if p:
            tokens.append(p)
    return tokens


def map_color_token(token: str, cfg: Dict[str, Any]) -> str:
    built_in = {
        "red": "红色",
    }
    built_in_color = built_in.get((token or "").strip().lower())
    if built_in_color:
        return built_in_color
    cmap = cfg.get("color_map", {})
    if token in cmap:
        return cmap[token]
    # 大小写兜底
    for k, v in cmap.items():
        if k.lower() == token.lower():
            return v
    return token


def output_attrs_by_color_mode(rule: Dict[str, Any], row: ParsedRow, cfg: Dict[str, Any]) -> List[str]:
    color, size = split_attr(row.attr)
    color_mode = rule.get("color_mode", "keep")

    if color_mode == "keep":
        return [row.attr]

    if color_mode == "split_attr_color":
        colors = [c.strip() for c in re.split(r"\+", color) if c.strip()]
        if not colors:
            return [row.attr]
        return [f"{c}-{size}" if size else c for c in colors]

    if color_mode == "fixed_colors":
        colors = normalize_separator_list(str(rule.get("fixed_colors", "")))
        if not colors:
            return [row.attr]
        return [f"{c}-{size}" if size else c for c in colors]

    if color_mode == "sku_color":
        tokens = english_color_tokens_from_sku(row.sku_code)
        colors = [map_color_token(t, cfg) for t in tokens]
        if not colors:
            return [row.attr]
        if any(re.search(r"[A-Za-z]", color) for color in colors):
            colors = [c.strip() for c in re.split(r"\+", color) if c.strip()]
            if not colors:
                return [row.attr]
        return [f"{c}-{size}" if size else c for c in colors]

    return [row.attr]


def rule_matches(rule: Dict[str, Any], sku_code: str) -> bool:
    if not rule.get("enabled", True):
        return False
    pattern = str(rule.get("pattern", ""))
    match_type = rule.get("match_type", "contains")
    if not pattern and match_type != "regex":
        return False
    try:
        patterns = [pattern] if match_type == "regex" else normalize_separator_list(pattern)
        if not patterns and pattern:
            patterns = [pattern]
        if match_type == "exact":
            return any(sku_code == p for p in patterns)
        if match_type == "startswith":
            return any(sku_code.startswith(p) for p in patterns)
        if match_type == "contains":
            return any(p in sku_code for p in patterns)
        if match_type == "regex":
            return re.search(pattern, sku_code) is not None
    except re.error:
        return False
    return False


def to_int(value: Any, default: int = 9999) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def sorted_indexed_rules(cfg: Dict[str, Any]) -> List[Tuple[int, Dict[str, Any]]]:
    return sorted(
        enumerate(cfg.get("rules", [])),
        key=lambda item: (to_int(item[1].get("priority", 9999)), item[0]),
    )


def sorted_indexed_group_rules(cfg: Dict[str, Any]) -> List[Tuple[int, Dict[str, Any]]]:
    return sorted(
        enumerate(cfg.get("group_rules", [])),
        key=lambda item: (to_int(item[1].get("priority", 9999)), item[0]),
    )


def sorted_rules(cfg: Dict[str, Any]) -> List[Dict[str, Any]]:
    return [rule for _, rule in sorted_indexed_rules(cfg)]


def sorted_group_rules(cfg: Dict[str, Any]) -> List[Dict[str, Any]]:
    return [rule for _, rule in sorted_indexed_group_rules(cfg)]


def resolve_group_name(style: str, matched_rule: Dict[str, Any], cfg: Dict[str, Any]) -> str:
    """按输出款号匹配分组规则；没有匹配时使用 SKU 规则中的分组名；仍为空则归为未分组。"""
    if is_quantity_only_style(style, cfg):
        return quantity_only_group_name(cfg)
    for gr in sorted_group_rules(cfg):
        if rule_matches(gr, style):
            name = str(gr.get("group_name", "")).strip()
            if name:
                return name
    rule_group = str(matched_rule.get("group_name", "")).strip()
    if rule_group:
        return rule_group
    return str(cfg.get("general", {}).get("default_group_name", "未分组") or "未分组")


def apply_rules_to_row(row: ParsedRow, cfg: Dict[str, Any]) -> List[OutputRow]:
    rules = sorted_rules(cfg)
    matched_rule: Optional[Dict[str, Any]] = None
    for rule in rules:
        if rule_matches(rule, row.sku_code):
            matched_rule = rule
            break

    if matched_rule is None:
        matched_rule = {
            "name": "无匹配规则，使用自动识别",
            "output_styles": "{auto}",
            "color_mode": "keep",
            "qty_multiplier": 1,
        }

    styles = parse_output_styles(str(matched_rule.get("output_styles", "{auto}")), row.sku_code, cfg)
    force_split_mg8_combo = should_split_mg8_combo(row, styles, cfg)
    attr_rule = matched_rule
    if force_split_mg8_combo:
        attr_rule = dict(matched_rule)
        attr_rule["color_mode"] = "split_attr_color"
    attrs = [normalize_output_attr(attr) for attr in output_attrs_by_color_mode(attr_rule, row, cfg)]
    try:
        multiplier = float(matched_rule.get("qty_multiplier", 1) or 1)
    except Exception:
        multiplier = 1
    dedupe_existing_attrs = force_split_mg8_combo or should_dedupe_existing_attrs(matched_rule)

    out_rows: List[OutputRow] = []
    for style in styles:
        style_attrs = [""] if is_attrless_quantity_style(style, cfg) else attrs
        for attr in style_attrs:
            qty = row.qty * multiplier
            # 一般报货数量应为整数；若用户设置了小数倍数，保留整数四舍五入并在明细中体现。
            if abs(qty - round(qty)) < 1e-9:
                qty_value = int(round(qty))
            else:
                qty_value = qty
            out_rows.append(OutputRow(
                source_file=row.source_file,
                page=row.page,
                original_attr=row.attr,
                sku_id=row.sku_id,
                sku_code=row.sku_code,
                output_style=style,
                group_name=resolve_group_name(style, matched_rule, cfg),
                output_attr=attr,
                qty=qty_value,
                rule_name=str(matched_rule.get("name", "未命名规则")),
                dedupe_existing_attrs=dedupe_existing_attrs,
            ))
    return out_rows


def attr_sort_key(attr: str) -> Tuple[int, str, int, str]:
    color, size = split_attr(attr)
    try:
        color_i = COLOR_ORDER.index(color)
    except ValueError:
        color_i = 999
    try:
        size_i = SIZE_ORDER.index(size)
    except ValueError:
        size_i = 999
    return color_i, color, size_i, size


def aggregate_output_rows(out_rows: List[OutputRow]) -> Tuple[OrderedDict, Dict[Tuple[str, str], int], Dict[str, int]]:
    """汇总为：分组 -> 款号 -> 属性集 -> 数量。"""
    groups: OrderedDict[str, OrderedDict[str, OrderedDict[str, float]]] = OrderedDict()
    style_first_index: Dict[Tuple[str, str], int] = {}
    group_first_index: Dict[str, int] = {}
    for idx, r in enumerate(out_rows):
        group = (r.group_name or "未分组").strip() or "未分组"
        if group not in groups:
            groups[group] = OrderedDict()
            group_first_index[group] = idx
        if r.output_style not in groups[group]:
            groups[group][r.output_style] = OrderedDict()
            style_first_index[(group, r.output_style)] = idx
        groups[group][r.output_style][r.output_attr] = groups[group][r.output_style].get(r.output_attr, 0) + r.qty
    return groups, style_first_index, group_first_index


def dedupe_split_attrs_against_existing(out_rows: List[OutputRow]) -> List[OutputRow]:
    existing_attrs = {
        (r.group_name, r.output_style, r.output_attr)
        for r in out_rows
        if not r.dedupe_existing_attrs
    }
    return [
        r for r in out_rows
        if not (r.dedupe_existing_attrs and (r.group_name, r.output_style, r.output_attr) in existing_attrs)
    ]


def group_sort_key(group: str, group_first_index: Dict[str, int], cfg: Optional[Dict[str, Any]] = None) -> Tuple[int, str]:
    if cfg is not None and group == quantity_only_group_name(cfg):
        return (-1, group)
    return (group_first_index.get(group, 999999), group)


def style_number_parts(style: str) -> Optional[Tuple[str, int, str]]:
    m = re.fullmatch(r"([A-Z]+(?:-[A-Z]+)*)-(\d+)", (style or "").strip())
    if not m:
        return None
    return m.group(1), int(m.group(2)), m.group(2)


def group_prefix(group: str) -> str:
    m = re.match(r"[A-Z]+(?:-[A-Z]+)*", (group or "").strip().upper())
    return m.group(0) if m else ""


def style_sort_key(group: str, style: str, first_index: Dict[Tuple[str, str], int]) -> Tuple[int, str, int, int, str]:
    first_seen = first_index.get((group, style), 999999)
    parts = style_number_parts(style)
    target_prefix = group_prefix(group)
    if parts and target_prefix and parts[0] == target_prefix:
        prefix, number, raw_number = parts
        return (0, prefix, number, len(raw_number), style)
    return (1, "", first_seen, 0, style)


def _fmt_qty(qty: float) -> Any:
    if isinstance(qty, float) and abs(qty - round(qty)) < 1e-9:
        return int(round(qty))
    return qty


def excel_text(value: str) -> str:
    return str(value).replace('"', '""')


def write_excel(
    groups: OrderedDict,
    first_index: Dict[Tuple[str, str], int],
    group_first_index: Dict[str, int],
    detail_rows: List[OutputRow],
    parsed_rows: List[ParsedRow],
    issues: List[ParseIssue],
    title: str,
    out_path: Path,
    cfg: Dict[str, Any],
) -> None:
    wb = Workbook()
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        pass
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    group_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="FFF2CC")

    ws = wb.active
    ws.title = "汇总"
    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(name="微软雅黑", size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for col, width in zip(["A", "B", "C", "D"], [16, 22, 10, 10]):
        ws.column_dimensions[col].width = width

    ws["A2"] = "不同款号"
    ws["B2"] = ""
    ws["C2"] = "总数量"
    ws["D2"] = ""
    for col in range(1, 5):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = total_fill
        cell.border = border

    row_idx = 4
    show_group_headers = bool(cfg.get("general", {}).get("show_group_headers", True))
    blank_between = bool(cfg.get("general", {}).get("blank_row_between_styles", True))
    group_header_rows: List[Tuple[int, str, bool]] = []

    for group in sorted(groups.keys(), key=lambda g: group_sort_key(g, group_first_index, cfg)):
        styles = groups[group]
        if show_group_headers:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            cell = ws.cell(row=row_idx, column=1)
            is_quantity_group = group == quantity_only_group_name(cfg)
            cell.value = group
            group_header_rows.append((row_idx, group, is_quantity_group))
            cell.font = Font(name="微软雅黑", size=11, bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = group_fill
            cell.border = border
            for col in range(2, 5):
                ws.cell(row=row_idx, column=col).border = border
                ws.cell(row=row_idx, column=col).fill = group_fill
            row_idx += 1

        for style in sorted(styles.keys(), key=lambda st: style_sort_key(group, st, first_index)):
            attrs = styles[style]
            items = sorted(attrs.items(), key=lambda kv: attr_sort_key(kv[0]))
            for n, (attr, qty) in enumerate(items):
                if n == 0:
                    ws.cell(row=row_idx, column=1, value=style)
                ws.cell(row=row_idx, column=2, value=attr)
                ws.cell(row=row_idx, column=3, value=_fmt_qty(qty))
                if n == len(items) - 1:
                    ws.cell(row=row_idx, column=4, value=f"=SUMIFS(C:C,H:H,H{row_idx},I:I,I{row_idx})")
                ws.cell(row=row_idx, column=5, value=1 if n == 0 else "")
                ws.cell(row=row_idx, column=6, value=group)
                ws.cell(row=row_idx, column=7, value=1 if n == 0 and not is_quantity_only_style(style, cfg) else "")
                ws.cell(row=row_idx, column=8, value=f'=IF(A{row_idx}<>"",A{row_idx},H{row_idx - 1})')
                ws.cell(row=row_idx, column=9, value=group)
                for col in range(1, 5):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = Font(name="微软雅黑", size=10, bold=(col == 1 and n == 0))
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                row_idx += 1
            if blank_between:
                ws.cell(row=row_idx, column=4, value=f'=IF(AND(A{row_idx}<>"",C{row_idx}<>""),SUMIFS(C:C,H:H,H{row_idx},I:I,I{row_idx}),"")')
                ws.cell(row=row_idx, column=8, value=f'=IF(A{row_idx}<>"",A{row_idx},H{row_idx - 1})')
                ws.cell(row=row_idx, column=9, value=group)
                row_idx += 1

    last_data_row = row_idx - 1
    if last_data_row < 4:
        last_data_row = 4
    data_end_formula = 'MATCH("合计",A:A,0)-1'
    ws["B2"] = (
        f'=COUNTIFS(A4:INDEX(A:A,{data_end_formula}),"<>",'
        f'C4:INDEX(C:C,{data_end_formula}),"<>")&"款"'
    )
    ws["D2"] = f"=SUM(C4:INDEX(C:C,{data_end_formula}))"
    for idx, (header_row, group, is_quantity_group) in enumerate(group_header_rows):
        group_name = excel_text(group)
        group_start_row = header_row + 1
        if idx + 1 < len(group_header_rows):
            group_end_row = group_header_rows[idx + 1][0] - 1
        else:
            group_end_row = last_data_row
        if group_end_row < group_start_row:
            group_end_row = group_start_row
        if is_quantity_group:
            ws.cell(row=header_row, column=1).value = (
                f'="{group_name}    数量："&SUM(C{group_start_row}:C{group_end_row})'
            )
        else:
            ws.cell(row=header_row, column=1).value = (
                f'="{group_name}    款数："&'
                f'COUNTIFS(A{group_start_row}:A{group_end_row},"<>",C{group_start_row}:C{group_end_row},"<>")'
                f'&"款    数量："&SUM(C{group_start_row}:C{group_end_row})'
            )

    ws.cell(row=row_idx, column=1, value="合计")
    ws.cell(
        row=row_idx,
        column=2,
        value=(
            f'=COUNTIFS(A4:INDEX(A:A,{data_end_formula}),"<>",'
            f'C4:INDEX(C:C,{data_end_formula}),"<>")&"款"'
        ),
    )
    ws.cell(row=row_idx, column=3, value=f"=SUM(C4:INDEX(C:C,{data_end_formula}))")
    for col in range(1, 5):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = total_fill
        cell.border = border
    ws.freeze_panes = "A4"
    for col in ["E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].hidden = True

    ds = wb.create_sheet("明细")
    headers = ["来源PDF", "页码", "原始SKU货号", "SKU ID", "原始属性集", "分组", "输出款号", "输出属性集", "数量", "命中规则"]
    ds.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ds.cell(row=1, column=c)
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.border = border
    for r in detail_rows:
        ds.append([r.source_file, r.page, r.sku_code, r.sku_id, r.original_attr, r.group_name, r.output_style, r.output_attr, r.qty, r.rule_name])
    widths = [24, 8, 36, 16, 18, 16, 14, 18, 8, 32]
    for idx, width in enumerate(widths, start=1):
        ds.column_dimensions[chr(64 + idx)].width = width
    for row_cells in ds.iter_rows(min_row=2):
        for cell in row_cells:
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    ds.freeze_panes = "A2"

    raw = wb.create_sheet("原始解析")
    raw_headers = ["来源PDF", "页码", "属性集", "SKU ID", "SKU货号", "实际发货数"]
    raw.append(raw_headers)
    for c in range(1, len(raw_headers) + 1):
        cell = raw.cell(row=1, column=c)
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for r in parsed_rows:
        raw.append([r.source_file, r.page, r.attr, r.sku_id, r.sku_code, r.qty])
    for idx, width in enumerate([24, 8, 18, 16, 36, 10], start=1):
        raw.column_dimensions[chr(64 + idx)].width = width
    for row_cells in raw.iter_rows(min_row=2):
        for cell in row_cells:
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    raw.freeze_panes = "A2"

    er = wb.create_sheet("异常检查")
    er_headers = ["来源PDF", "页码", "异常类型", "内容"]
    er.append(er_headers)
    for c in range(1, len(er_headers) + 1):
        cell = er.cell(row=1, column=c)
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE4D6")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    if issues:
        for issue in issues:
            er.append([issue.source_file, issue.page or "", issue.issue, issue.text])
    else:
        er.append(["", "", "无异常", "解析过程没有发现明显异常"])
    for idx, width in enumerate([24, 8, 24, 80], start=1):
        er.column_dimensions[chr(64 + idx)].width = width
    for row_cells in er.iter_rows(min_row=2):
        for cell in row_cells:
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    er.freeze_panes = "A2"

    rs = wb.create_sheet("规则快照")
    rs.append(["配置文件版本", cfg.get("version", "")])
    rs.append([])
    rs.append(["启用", "优先级", "规则名", "分组", "匹配方式", "匹配内容", "输出款号", "颜色模式", "固定颜色", "数量倍数", "备注"])
    for rule in sorted_rules(cfg):
        rs.append([
            "是" if rule.get("enabled", True) else "否",
            rule.get("priority", ""),
            rule.get("name", ""),
            rule.get("group_name", ""),
            rule.get("match_type", ""),
            rule.get("pattern", ""),
            rule.get("output_styles", ""),
            rule.get("color_mode", ""),
            rule.get("fixed_colors", ""),
            rule.get("qty_multiplier", ""),
            rule.get("note", ""),
        ])
    for row_cells in rs.iter_rows():
        for cell in row_cells:
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    for idx, width in enumerate([8, 8, 36, 16, 12, 32, 24, 18, 24, 10, 60], start=1):
        rs.column_dimensions[chr(64 + idx)].width = width
    rs.freeze_panes = "A4"

    grs = wb.create_sheet("分组规则快照")
    grs.append(["启用", "优先级", "规则名", "匹配方式", "匹配内容", "分组名称"])
    for gr in sorted_group_rules(cfg):
        grs.append([
            "是" if gr.get("enabled", True) else "否",
            gr.get("priority", ""),
            gr.get("name", ""),
            gr.get("match_type", ""),
            gr.get("pattern", ""),
            gr.get("group_name", ""),
        ])
    for row_cells in grs.iter_rows():
        for cell in row_cells:
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    for idx, width in enumerate([8, 8, 36, 12, 32, 18], start=1):
        grs.column_dimensions[chr(64 + idx)].width = width
    grs.freeze_panes = "A2"

    wb.save(out_path)


def collect_pdf_paths(inputs: Iterable[Path]) -> List[Path]:
    result: List[Path] = []
    for p in inputs:
        if p.is_file() and p.suffix.lower() == ".pdf":
            result.append(p)
        elif p.is_dir():
            result.extend(sorted(x for x in p.iterdir() if x.is_file() and x.suffix.lower() == ".pdf"))
    # 去重但保序
    seen = set()
    out = []
    for p in result:
        rp = str(p.resolve())
        if rp not in seen:
            seen.add(rp)
            out.append(p)
    return out


def process_pdfs(pdf_paths: List[Path], out_path: Path, cfg: Dict[str, Any]) -> Dict[str, Any]:
    all_parsed: List[ParsedRow] = []
    all_issues: List[ParseIssue] = []
    titles: List[str] = []

    for pdf in pdf_paths:
        rows, issues, title = parse_pdf_rows(pdf)
        all_parsed.extend(rows)
        all_issues.extend(issues)
        titles.append(title)

    detail_rows: List[OutputRow] = []
    for row in all_parsed:
        detail_rows.extend(apply_rules_to_row(row, cfg))
    detail_rows = dedupe_split_attrs_against_existing(detail_rows)

    groups, first_index, group_first_index = aggregate_output_rows(detail_rows)
    title = titles[0] if titles else "报货汇总"
    if len(set(titles)) > 1:
        title = "报货汇总"

    write_excel(groups, first_index, group_first_index, detail_rows, all_parsed, all_issues, title, out_path, cfg)

    original_qty = sum(r.qty for r in all_parsed)
    output_qty = sum(sum(sum(attrs.values()) for attrs in styles.values()) for styles in groups.values())
    style_count = len({style for styles in groups.values() for style in styles.keys()})
    summary_line_count = sum(len(attrs) for styles in groups.values() for attrs in styles.values())
    return {
        "out_path": str(out_path),
        "pdf_count": len(pdf_paths),
        "parsed_rows": len(all_parsed),
        "original_qty": original_qty,
        "output_lines": style_count,
        "summary_line_count": summary_line_count,
        "output_qty": output_qty,
        "issues": len(all_issues),
    }


# ============================== GUI ==============================

def enable_dpi_awareness() -> None:
    """让 Windows 高缩放屏幕使用真实 DPI，避免界面被系统拉伸后发糊。"""
    if sys.platform != "win32":
        return
    try:
        import ctypes

        try:
            ctypes.windll.user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-4))
            return
        except Exception:
            pass
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(2)
            return
        except Exception:
            pass
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

if tk is not None:
    class RuleDialog(tk.Toplevel):
        def __init__(self, master, rule: Optional[Dict[str, Any]] = None):
            super().__init__(master)
            self.title("规则设置")
            self.resizable(False, False)
            self.result: Optional[Dict[str, Any]] = None
            self.rule = rule or {
                "enabled": True,
                "name": "新规则",
                "priority": 100,
                "match_type": "contains",
                "pattern": "",
                "output_styles": "{auto}",
                "color_mode": "keep",
                "fixed_colors": "",
                "qty_multiplier": 1,
                "group_name": "",
                "note": "",
            }
            self.vars: Dict[str, tk.Variable] = {}
            self._build()
            self.grab_set()
            self.transient(master)

        def _build(self):
            pad = {"padx": 8, "pady": 5}
            frm = ttk.Frame(self)
            frm.grid(row=0, column=0, sticky="nsew")

            self.vars["enabled"] = tk.BooleanVar(value=bool(self.rule.get("enabled", True)))
            ttk.Checkbutton(frm, text="启用此规则", variable=self.vars["enabled"]).grid(row=0, column=0, columnspan=2, sticky="w", **pad)

            fields = [
                ("规则名", "name", "entry"),
                ("优先级", "priority", "entry"),
                ("匹配方式", "match_type", "combo_match"),
                ("匹配内容", "pattern", "entry"),
                ("输出款号", "output_styles", "entry"),
                ("颜色模式", "color_mode", "combo_color"),
                ("固定颜色", "fixed_colors", "entry"),
                ("数量倍数", "qty_multiplier", "entry"),
                ("默认分组", "group_name", "entry"),
                ("备注", "note", "entry"),
            ]
            row = 1
            for label, key, kind in fields:
                ttk.Label(frm, text=label).grid(row=row, column=0, sticky="e", **pad)
                if kind == "combo_match":
                    var = tk.StringVar(value=str(self.rule.get(key, "contains")))
                    self.vars[key] = var
                    cb = ttk.Combobox(frm, textvariable=var, width=38, state="readonly",
                                      values=["contains", "exact", "startswith", "regex"])
                    cb.grid(row=row, column=1, sticky="w", **pad)
                elif kind == "combo_color":
                    var = tk.StringVar(value=str(self.rule.get(key, "keep")))
                    self.vars[key] = var
                    cb = ttk.Combobox(frm, textvariable=var, width=38, state="readonly",
                                      values=["keep", "split_attr_color", "fixed_colors", "sku_color"])
                    cb.grid(row=row, column=1, sticky="w", **pad)
                else:
                    var = tk.StringVar(value=str(self.rule.get(key, "")))
                    self.vars[key] = var
                    ttk.Entry(frm, textvariable=var, width=42).grid(row=row, column=1, sticky="w", **pad)
                row += 1

            help_text = (
                "说明：\n"
                "1. 输出款号写 {auto} 表示从 SKU货号 自动识别。\n"
                "2. 拆两个款式时写：MG-55,SSL-14。\n"
                "3. keep=保留原属性集；split_attr_color=把 白色+卡其色-L 拆成 白色-L/卡其色-L。\n"
                "4. sku_color=从 SKU 英文颜色拆色，例如 White+Khaki。\n"
                "5. fixed_colors=使用你手动填写的颜色列表，例如 白色,卡其色。\n"
                "6. 默认分组只在没有命中‘分组设置’时生效；分组设置优先级更高。"
            )
            ttk.Label(frm, text=help_text, justify="left", foreground="#555").grid(row=row, column=0, columnspan=2, sticky="w", **pad)
            row += 1

            btns = ttk.Frame(frm)
            btns.grid(row=row, column=0, columnspan=2, pady=8)
            ttk.Button(btns, text="保存", command=self._save).pack(side="left", padx=6)
            ttk.Button(btns, text="取消", command=self.destroy).pack(side="left", padx=6)

        def _save(self):
            try:
                priority = int(str(self.vars["priority"].get()).strip() or "100")
            except ValueError:
                messagebox.showerror("错误", "优先级必须是整数。")
                return
            try:
                multiplier = float(str(self.vars["qty_multiplier"].get()).strip() or "1")
            except ValueError:
                messagebox.showerror("错误", "数量倍数必须是数字。")
                return
            self.result = {
                "enabled": bool(self.vars["enabled"].get()),
                "name": self.vars["name"].get().strip() or "未命名规则",
                "priority": priority,
                "match_type": self.vars["match_type"].get(),
                "pattern": self.vars["pattern"].get().strip(),
                "output_styles": self.vars["output_styles"].get().strip() or "{auto}",
                "color_mode": self.vars["color_mode"].get(),
                "fixed_colors": self.vars["fixed_colors"].get().strip(),
                "qty_multiplier": multiplier,
                "group_name": self.vars["group_name"].get().strip(),
                "note": self.vars["note"].get().strip(),
            }
            self.destroy()


    class GroupRuleDialog(tk.Toplevel):
        def __init__(self, master, rule: Optional[Dict[str, Any]] = None):
            super().__init__(master)
            self.title("分组规则设置")
            self.resizable(False, False)
            self.result: Optional[Dict[str, Any]] = None
            self.rule = rule or {
                "enabled": True,
                "priority": 100,
                "name": "新分组规则",
                "match_type": "startswith",
                "pattern": "",
                "group_name": "",
            }
            self.vars: Dict[str, tk.Variable] = {}
            self._build()
            self.grab_set()
            self.transient(master)

        def _build(self):
            pad = {"padx": 8, "pady": 5}
            frm = ttk.Frame(self)
            frm.grid(row=0, column=0, sticky="nsew")

            self.vars["enabled"] = tk.BooleanVar(value=bool(self.rule.get("enabled", True)))
            ttk.Checkbutton(frm, text="启用此分组规则", variable=self.vars["enabled"]).grid(row=0, column=0, columnspan=2, sticky="w", **pad)

            fields = [
                ("规则名", "name", "entry"),
                ("优先级", "priority", "entry"),
                ("匹配方式", "match_type", "combo_match"),
                ("匹配款号", "pattern", "entry"),
                ("分组名称", "group_name", "entry"),
            ]
            row = 1
            for label, key, kind in fields:
                ttk.Label(frm, text=label).grid(row=row, column=0, sticky="e", **pad)
                if kind == "combo_match":
                    var = tk.StringVar(value=str(self.rule.get(key, "startswith")))
                    self.vars[key] = var
                    cb = ttk.Combobox(frm, textvariable=var, width=38, state="readonly",
                                      values=["contains", "exact", "startswith", "regex"])
                    cb.grid(row=row, column=1, sticky="w", **pad)
                else:
                    var = tk.StringVar(value=str(self.rule.get(key, "")))
                    self.vars[key] = var
                    ttk.Entry(frm, textvariable=var, width=42).grid(row=row, column=1, sticky="w", **pad)
                row += 1

            help_text = (
                "说明：分组规则按‘输出款号’匹配，不按 PDF 原始 SKU 匹配。\n"
                "例：匹配方式 startswith，匹配款号 MG-，分组名称 MG系列，所有 MG- 开头的款会排在一起。\n"
                "多个匹配款号可用逗号分隔，例如 CW-,YYCK- 会归到同一个分组。\n"
                "特殊指定用 exact，例如 SK-010,SK-011 分组名称 MG，会把这些款直接归到 MG。\n"
                "如果同一个款号命中多个分组规则，优先级数字越小越先使用。"
            )
            ttk.Label(frm, text=help_text, justify="left", foreground="#555").grid(row=row, column=0, columnspan=2, sticky="w", **pad)
            row += 1

            btns = ttk.Frame(frm)
            btns.grid(row=row, column=0, columnspan=2, pady=8)
            ttk.Button(btns, text="保存", command=self._save).pack(side="left", padx=6)
            ttk.Button(btns, text="取消", command=self.destroy).pack(side="left", padx=6)

        def _save(self):
            try:
                priority = int(str(self.vars["priority"].get()).strip() or "100")
            except ValueError:
                messagebox.showerror("错误", "优先级必须是整数。")
                return
            group_name = self.vars["group_name"].get().strip()
            if not group_name:
                messagebox.showerror("错误", "分组名称不能为空。")
                return
            self.result = {
                "enabled": bool(self.vars["enabled"].get()),
                "name": self.vars["name"].get().strip() or "未命名分组规则",
                "priority": priority,
                "match_type": self.vars["match_type"].get(),
                "pattern": self.vars["pattern"].get().strip(),
                "group_name": group_name,
            }
            self.destroy()


    class BaohuoApp(tk.Tk):
        def __init__(self):
            enable_dpi_awareness()
            super().__init__()
            self.title(APP_NAME)
            self.cfg_path = default_config_path()
            self.cfg = load_config(self.cfg_path)
            self.pdf_paths: List[Path] = []
            self.processing = False
            self.last_output_path: Optional[Path] = None
            self._configure_window_metrics()
            self._apply_theme()
            self._build_ui()
            self.refresh_rules_tree()
            self.refresh_group_rules_tree()
            self.refresh_color_table()
            self.refresh_prefix_table()
            self.refresh_style_code_table()
            self.refresh_quantity_only_table()
            self.refresh_pdf_tree()

        def _configure_window_metrics(self):
            screen_w = max(self.winfo_screenwidth(), 900)
            screen_h = max(self.winfo_screenheight(), 640)
            usable_w = max(screen_w - 80, 900)
            usable_h = max(screen_h - 110, 620)

            target_w = min(max(int(screen_w * 0.78), 1120), 1560, usable_w)
            target_h = min(max(int(screen_h * 0.78), 720), 980, usable_h)
            min_w = min(1020, usable_w)
            min_h = min(660, usable_h)

            x = max((screen_w - target_w) // 2, 0)
            y = max((screen_h - target_h) // 2, 0)
            self.geometry(f"{target_w}x{target_h}+{x}+{y}")
            self.minsize(min_w, min_h)
            self.compact_ui = target_w < 1160 or target_h < 720

            try:
                dpi_scale = float(self.winfo_fpixels("1i")) / 72.0
                self.tk.call("tk", "scaling", max(1.0, min(dpi_scale, 2.4)))
            except Exception:
                pass

        def _choose_font_family(self) -> str:
            preferred = ["PingFang SC", "Microsoft YaHei UI", "Microsoft YaHei", "微软雅黑"]
            if tkfont is not None:
                families = {name.lower(): name for name in tkfont.families(self)}
                for name in preferred:
                    if name.lower() in families:
                        return families[name.lower()]
            return "Microsoft YaHei UI"

        def _apply_theme(self):
            self.colors = {
                "bg": "#111318",
                "panel": "#181B22",
                "panel_2": "#20242D",
                "control": "#252A34",
                "control_hover": "#303744",
                "control_active": "#384150",
                "input_bg": "#141922",
                "input_focus": "#1A202B",
                "text": "#F2F4F8",
                "muted": "#AAB1C0",
                "accent": "#D6B46A",
                "accent_hover": "#E2C47B",
                "accent_active": "#C9A75E",
                "border": "#303643",
                "disabled": "#6F7785",
                "danger": "#E06C75",
                "log_text": "#D7DBE5",
            }
            c = self.colors
            self.configure(bg=c["bg"])
            self.style = ttk.Style(self)
            try:
                self.style.theme_use("clam")
            except tk.TclError:
                pass
            self.font_family = self._choose_font_family()
            base_size = 9 if self.compact_ui else 10
            title_size = 16 if self.compact_ui else 18
            small_size = 8 if self.compact_ui else 9
            row_height = 26 if self.compact_ui else 28
            base_font = (self.font_family, base_size)
            title_font = (self.font_family, title_size, "bold")
            small_font = (self.font_family, small_size)
            self.option_add("*Font", base_font)
            self.option_add("*TCombobox*Listbox.background", c["control"])
            self.option_add("*TCombobox*Listbox.foreground", c["text"])
            self.option_add("*TCombobox*Listbox.selectBackground", c["accent"])
            self.option_add("*TCombobox*Listbox.selectForeground", c["bg"])
            self.option_add("*TCombobox*Listbox.borderWidth", 0)
            self.style.configure(".", font=base_font, background=c["bg"], foreground=c["text"])
            self.style.configure("App.TFrame", background=c["bg"])
            self.style.configure("Panel.TFrame", background=c["panel"])
            self.style.configure("Toolbar.TFrame", background=c["panel"])
            self.style.configure("Title.TLabel", font=title_font, background=c["bg"], foreground=c["text"])
            self.style.configure("Subtitle.TLabel", font=small_font, background=c["bg"], foreground=c["muted"])
            self.style.configure("PanelTitle.TLabel", font=(self.font_family, base_size + 1, "bold"), background=c["panel"], foreground=c["accent"])
            self.style.configure("Muted.TLabel", font=small_font, background=c["panel"], foreground=c["muted"])
            self.style.configure("Status.TLabel", font=small_font, background=c["panel_2"], foreground=c["accent"])
            self.style.configure("TButton", padding=((10 if self.compact_ui else 12), (6 if self.compact_ui else 7)), relief="flat", background=c["control"], foreground=c["text"], bordercolor=c["control"], darkcolor=c["control"], lightcolor=c["control"], focuscolor=c["control"])
            self.style.map("TButton", background=[("disabled", c["panel_2"]), ("pressed", c["control_active"]), ("active", c["control_hover"])], foreground=[("disabled", c["disabled"])], bordercolor=[("pressed", c["control_active"]), ("active", c["control_hover"])])
            self.style.configure("Accent.TButton", padding=((13 if self.compact_ui else 16), (7 if self.compact_ui else 8)), background=c["accent"], foreground=c["bg"], bordercolor=c["accent"], darkcolor=c["accent"], lightcolor=c["accent"], focuscolor=c["accent"])
            self.style.map("Accent.TButton", background=[("disabled", "#6F654C"), ("pressed", c["accent_active"]), ("active", c["accent_hover"])], foreground=[("disabled", "#2A261D")], bordercolor=[("pressed", c["accent_active"]), ("active", c["accent_hover"])])
            self.style.configure("Danger.TButton", padding=((10 if self.compact_ui else 12), (6 if self.compact_ui else 7)), foreground=c["danger"], background=c["control"], bordercolor=c["control"])
            self.style.configure("Nav.TButton", padding=(10, 9), anchor="w", background=c["panel"], foreground=c["muted"], bordercolor=c["panel"], darkcolor=c["panel"], lightcolor=c["panel"])
            self.style.map("Nav.TButton", background=[("active", c["panel_2"])], foreground=[("active", c["text"])])
            self.style.configure("ActiveNav.TButton", padding=(10, 9), anchor="w", background=c["panel_2"], foreground=c["accent"], bordercolor=c["panel_2"], darkcolor=c["panel_2"], lightcolor=c["panel_2"])
            self.style.map("ActiveNav.TButton", background=[("active", c["panel_2"])], foreground=[("active", c["accent"])])
            self.style.configure("TEntry", fieldbackground=c["input_bg"], background=c["input_bg"], foreground=c["text"], insertcolor=c["text"], bordercolor=c["border"], lightcolor=c["border"], darkcolor=c["border"], relief="flat")
            self.style.map("TEntry", fieldbackground=[("focus", c["input_focus"])], bordercolor=[("focus", c["accent"])], lightcolor=[("focus", c["accent"])], darkcolor=[("focus", c["accent"])])
            self.style.configure("TCombobox", fieldbackground=c["control"], background=c["control"], foreground=c["text"], selectforeground=c["text"], selectbackground=c["control"], arrowcolor=c["muted"], bordercolor=c["control"], darkcolor=c["control"], lightcolor=c["control"], relief="flat")
            self.style.map("TCombobox", fieldbackground=[("readonly", c["control"]), ("focus", c["control_hover"]), ("active", c["control_hover"])], background=[("readonly", c["control"]), ("pressed", c["control_active"]), ("active", c["control_hover"])], foreground=[("readonly", c["text"])], selectforeground=[("readonly", c["text"])], selectbackground=[("readonly", c["control"])], arrowcolor=[("active", c["text"]), ("readonly", c["muted"])], bordercolor=[("focus", c["accent"]), ("active", c["control_hover"])])
            self.style.configure("Treeview", font=base_font, rowheight=row_height, background=c["panel"], fieldbackground=c["panel"], foreground=c["text"], borderwidth=0)
            self.style.configure("Treeview.Heading", font=(self.font_family, base_size, "bold"), padding=(6, 7), background=c["panel_2"], foreground=c["accent"], bordercolor=c["border"])
            self.style.map("Treeview", background=[("selected", c["accent"])], foreground=[("selected", c["bg"])])
            self.style.configure("Vertical.TScrollbar", background=c["control"], darkcolor=c["control"], lightcolor=c["control"], troughcolor=c["bg"], bordercolor=c["bg"], arrowcolor=c["muted"], relief="flat", arrowsize=12)
            self.style.configure("Horizontal.TScrollbar", background=c["control"], darkcolor=c["control"], lightcolor=c["control"], troughcolor=c["bg"], bordercolor=c["bg"], arrowcolor=c["muted"], relief="flat", arrowsize=12)

        def _build_ui(self):
            page_pad = 12 if self.compact_ui else 18
            header_y = (10, 6) if self.compact_ui else (16, 8)
            bottom_y = (0, 10) if self.compact_ui else (0, 14)
            self.status_var = tk.StringVar(value="就绪")
            self.pdf_count_var = tk.StringVar(value="0 个 PDF")
            self.result_summary_var = tk.StringVar(value="尚未处理")
            self.config_state_var = tk.StringVar(value=f"配置：{self.cfg_path.name}")

            root = ttk.Frame(self, style="App.TFrame")
            root.pack(fill="both", expand=True)

            header = ttk.Frame(root, style="App.TFrame")
            header.pack(fill="x", padx=page_pad, pady=header_y)
            ttk.Label(header, text=APP_NAME, style="Title.TLabel").pack(side="left")
            ttk.Label(header, textvariable=self.config_state_var, style="Subtitle.TLabel").pack(side="right", pady=(8, 0))

            main = ttk.Frame(root, style="App.TFrame")
            main.pack(fill="both", expand=True, padx=page_pad, pady=(0, 10 if self.compact_ui else 12))
            nav_width = 132 if self.compact_ui else 156
            nav = ttk.Frame(main, style="Panel.TFrame", padding=8)
            nav.pack(side="left", fill="y", padx=(0, 10 if self.compact_ui else 12))
            nav.pack_propagate(False)
            nav.configure(width=nav_width)

            self.content_stack = ttk.Frame(main, style="App.TFrame")
            self.content_stack.pack(side="left", fill="both", expand=True)
            self.content_stack.rowconfigure(0, weight=1)
            self.content_stack.columnconfigure(0, weight=1)

            self.nav_buttons: Dict[str, ttk.Button] = {}
            nav_items = [
                ("process", "处理PDF"),
                ("rules", "SKU规则"),
                ("groups", "分组设置"),
                ("color", "颜色/前缀"),
            ]
            for key, label in nav_items:
                btn = ttk.Button(nav, text=label, style="Nav.TButton", command=lambda k=key: self._show_page(k))
                btn.pack(fill="x", pady=(0, 6))
                self.nav_buttons[key] = btn

            self.tab_process = ttk.Frame(self.content_stack, style="App.TFrame")
            self.tab_rules = ttk.Frame(self.content_stack, style="App.TFrame")
            self.tab_groups = ttk.Frame(self.content_stack, style="App.TFrame")
            self.tab_color = ttk.Frame(self.content_stack, style="App.TFrame")
            self.pages = {
                "process": self.tab_process,
                "rules": self.tab_rules,
                "groups": self.tab_groups,
                "color": self.tab_color,
            }
            for page in self.pages.values():
                page.grid(row=0, column=0, sticky="nsew")
            self._build_process_tab()
            self._build_rules_tab()
            self._build_group_tab()
            self._build_color_tab()
            self._show_page("process")

            status = ttk.Frame(root, style="App.TFrame")
            status.pack(fill="x", padx=page_pad, pady=bottom_y)
            ttk.Label(status, textvariable=self.status_var, style="Status.TLabel", anchor="w").pack(fill="x", ipady=5)

        def _panel(self, parent) -> ttk.Frame:
            panel = ttk.Frame(parent, style="Panel.TFrame", padding=8 if self.compact_ui else 12)
            return panel

        def _show_page(self, key: str):
            page = self.pages.get(key)
            if page is None:
                return
            page.tkraise()
            for nav_key, btn in self.nav_buttons.items():
                btn.configure(style="ActiveNav.TButton" if nav_key == key else "Nav.TButton")

        def _set_status(self, text: str):
            self.status_var.set(text)
            self.update_idletasks()

        def _mark_config_dirty(self):
            self.config_state_var.set(f"配置：{self.cfg_path.name}（未保存）")

        def _default_output_path(self) -> Path:
            output_name = str(self.cfg.get("general", {}).get("default_output_name", "报货汇总.xlsx") or "报货汇总.xlsx")
            base = self.pdf_paths[0].parent if self.pdf_paths else Path.cwd()
            return (base / output_name).resolve()

        def _set_busy(self, busy: bool):
            self.processing = busy
            state = "disabled" if busy else "normal"
            for btn in (getattr(self, "run_btn", None), getattr(self, "add_pdf_btn", None), getattr(self, "add_folder_btn", None)):
                if btn is not None:
                    btn.configure(state=state)

        def _build_process_tab(self):
            container = ttk.Frame(self.tab_process, style="App.TFrame", padding=(0, 12, 0, 0))
            container.pack(fill="both", expand=True)

            toolbar = self._panel(container)
            toolbar.pack(fill="x", padx=0, pady=(0, 10))
            ttk.Label(toolbar, text="PDF 队列", style="PanelTitle.TLabel").pack(side="left", padx=(0, 12))
            ttk.Label(toolbar, textvariable=self.pdf_count_var, style="Muted.TLabel").pack(side="left")
            self.add_pdf_btn = ttk.Button(toolbar, text="添加PDF", command=self.add_pdfs)
            self.add_pdf_btn.pack(side="right", padx=(6, 0))
            self.add_folder_btn = ttk.Button(toolbar, text="添加文件夹", command=self.add_folder)
            self.add_folder_btn.pack(side="right", padx=(6, 0))
            self.remove_pdf_btn = ttk.Button(toolbar, text="移除选中", command=self.remove_selected_pdf)
            self.remove_pdf_btn.pack(side="right", padx=(6, 0))
            self.clear_pdf_btn = ttk.Button(toolbar, text="清空", command=self.clear_pdfs, style="Danger.TButton")
            self.clear_pdf_btn.pack(side="right", padx=(6, 0))

            list_panel = self._panel(container)
            list_panel.pack(fill="both", expand=True, padx=0, pady=(0, 10))
            cols = ("name", "folder", "path")
            self.pdf_tree = ttk.Treeview(list_panel, columns=cols, show="headings", height=11, selectmode="extended")
            self.pdf_tree.heading("name", text="文件名")
            self.pdf_tree.heading("folder", text="所在目录")
            self.pdf_tree.heading("path", text="")
            self.pdf_tree.column("name", width=260, anchor="w")
            self.pdf_tree.column("folder", width=760, anchor="w")
            self.pdf_tree.column("path", width=0, minwidth=0, stretch=False)
            self.pdf_tree.pack(side="left", fill="both", expand=True)
            sb = ttk.Scrollbar(list_panel, orient="vertical", command=self.pdf_tree.yview)
            self.pdf_tree.configure(yscrollcommand=sb.set)
            sb.pack(side="right", fill="y")
            self.pdf_tree.bind("<Delete>", lambda e: self.remove_selected_pdf())

            output_panel = self._panel(container)
            output_panel.pack(fill="x", padx=0, pady=(0, 10))
            ttk.Label(output_panel, text="输出 Excel", style="PanelTitle.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
            self.output_var = tk.StringVar(value=str(self._default_output_path()))
            output_entry = ttk.Entry(output_panel, textvariable=self.output_var)
            output_entry.grid(row=1, column=0, sticky="ew", padx=(0, 8))
            ttk.Button(output_panel, text="选择位置", command=self.choose_output).grid(row=1, column=1, padx=(0, 8))
            self.run_btn = ttk.Button(output_panel, text="开始处理", command=self.run_process, style="Accent.TButton")
            self.run_btn.grid(row=1, column=2)
            self.open_output_btn = ttk.Button(output_panel, text="打开输出目录", command=self.open_output_folder, state="disabled")
            self.open_output_btn.grid(row=1, column=3, padx=(8, 0))
            ttk.Label(output_panel, textvariable=self.result_summary_var, style="Muted.TLabel").grid(row=2, column=0, columnspan=4, sticky="w", pady=(8, 0))
            output_panel.columnconfigure(0, weight=1)

            log_panel = self._panel(container)
            log_panel.pack(fill="both", expand=True)
            log_header = ttk.Frame(log_panel, style="Panel.TFrame")
            log_header.pack(fill="x", pady=(0, 8))
            ttk.Label(log_header, text="运行日志", style="PanelTitle.TLabel").pack(side="left")
            ttk.Button(log_header, text="清空日志", command=self.clear_log).pack(side="right")
            c = self.colors
            self.log = tk.Text(log_panel, height=8, bg=c["input_bg"], fg=c["log_text"], insertbackground=c["log_text"], relief="flat", highlightthickness=1, highlightbackground=c["border"], highlightcolor=c["accent"], wrap="word")
            self.log.pack(side="left", fill="both", expand=True)
            log_sb = ttk.Scrollbar(log_panel, orient="vertical", command=self.log.yview)
            self.log.configure(yscrollcommand=log_sb.set)
            log_sb.pack(side="right", fill="y")

        def _build_rules_tab(self):
            container = ttk.Frame(self.tab_rules, style="App.TFrame", padding=(0, 12, 0, 0))
            container.pack(fill="both", expand=True)
            btnfrm = self._panel(container)
            btnfrm.pack(fill="x", pady=(0, 10))
            ttk.Label(btnfrm, text="SKU / 套装规则", style="PanelTitle.TLabel").pack(side="left", padx=(0, 12))
            ttk.Button(btnfrm, text="新增", command=self.add_rule).pack(side="left", padx=(0, 6))
            ttk.Button(btnfrm, text="编辑", command=self.edit_rule).pack(side="left", padx=(0, 6))
            ttk.Button(btnfrm, text="删除", command=self.delete_rule, style="Danger.TButton").pack(side="left", padx=(0, 6))
            ttk.Button(btnfrm, text="恢复默认", command=self.reset_config).pack(side="right", padx=(6, 0))
            ttk.Button(btnfrm, text="导出配置", command=self.export_config).pack(side="right", padx=(6, 0))
            ttk.Button(btnfrm, text="导入配置", command=self.import_config).pack(side="right", padx=(6, 0))
            ttk.Button(btnfrm, text="保存配置", command=self.save_all_config, style="Accent.TButton").pack(side="right", padx=(6, 0))

            body = ttk.PanedWindow(container, orient="horizontal")
            body.pack(fill="both", expand=True)

            list_panel = self._panel(body)
            detail_panel = self._panel(body)
            body.add(list_panel, weight=2)
            body.add(detail_panel, weight=3)

            ttk.Label(list_panel, text="规则列表", style="PanelTitle.TLabel").pack(anchor="w", pady=(0, 8))
            cols = ("state", "priority", "name")
            self.rules_tree = ttk.Treeview(list_panel, columns=cols, show="headings", selectmode="browse")
            headings = {"state": "状态", "priority": "优先级", "name": "规则名"}
            widths = {"state": 58, "priority": 70, "name": 360 if not self.compact_ui else 280}
            for col in cols:
                self.rules_tree.heading(col, text=headings[col])
                self.rules_tree.column(col, width=widths[col], anchor="center" if col != "name" else "w", stretch=(col == "name"))
            self.rules_tree.pack(side="left", fill="both", expand=True)
            rules_sb = ttk.Scrollbar(list_panel, orient="vertical", command=self.rules_tree.yview)
            self.rules_tree.configure(yscrollcommand=rules_sb.set)
            rules_sb.pack(side="right", fill="y")
            self.rules_tree.bind("<Double-1>", lambda e: self.edit_rule())
            self.rules_tree.bind("<Delete>", lambda e: self.delete_rule())
            self.rules_tree.bind("<<TreeviewSelect>>", lambda e: self.refresh_rule_detail())

            ttk.Label(detail_panel, text="规则详情", style="PanelTitle.TLabel").pack(anchor="w", pady=(0, 8))
            self.rule_detail_title_var = tk.StringVar(value="请选择一条规则")
            self.rule_detail_meta_var = tk.StringVar(value="")
            ttk.Label(detail_panel, textvariable=self.rule_detail_title_var, style="PanelTitle.TLabel", wraplength=520, justify="left").pack(anchor="w")
            ttk.Label(detail_panel, textvariable=self.rule_detail_meta_var, style="Muted.TLabel", wraplength=520, justify="left").pack(anchor="w", pady=(3, 12))
            self.rule_detail_fields: Dict[str, tk.StringVar] = {}
            detail_fields = [
                ("匹配方式", "match_type"),
                ("匹配内容", "pattern"),
                ("输出款号", "output_styles"),
                ("颜色模式", "color_mode"),
                ("固定颜色", "fixed_colors"),
                ("默认分组", "group_name"),
                ("数量倍数", "qty_multiplier"),
                ("备注", "note"),
            ]
            grid = ttk.Frame(detail_panel, style="Panel.TFrame")
            grid.pack(fill="both", expand=True)
            for row, (label, key) in enumerate(detail_fields):
                ttk.Label(grid, text=label, style="Muted.TLabel").grid(row=row, column=0, sticky="ne", padx=(0, 12), pady=5)
                var = tk.StringVar(value="")
                self.rule_detail_fields[key] = var
                ttk.Label(grid, textvariable=var, background=self.colors["panel"], foreground=self.colors["text"], wraplength=620, justify="left").grid(row=row, column=1, sticky="nw", pady=5)
            grid.columnconfigure(1, weight=1)

        def _build_group_tab(self):
            container = ttk.Frame(self.tab_groups, style="App.TFrame", padding=(0, 12, 0, 0))
            container.pack(fill="both", expand=True)
            btnfrm = self._panel(container)
            btnfrm.pack(fill="x", pady=(0, 10))
            ttk.Label(btnfrm, text="分组设置", style="PanelTitle.TLabel").pack(side="left", padx=(0, 12))
            ttk.Button(btnfrm, text="新增", command=self.add_group_rule).pack(side="left", padx=(0, 6))
            ttk.Button(btnfrm, text="新增特殊规则", command=self.add_special_group_rule).pack(side="left", padx=(0, 6))
            ttk.Button(btnfrm, text="编辑", command=self.edit_group_rule).pack(side="left", padx=(0, 6))
            ttk.Button(btnfrm, text="删除", command=self.delete_group_rule, style="Danger.TButton").pack(side="left", padx=(0, 6))
            ttk.Button(btnfrm, text="保存配置", command=self.save_all_config, style="Accent.TButton").pack(side="right")

            tip_panel = self._panel(container)
            tip_panel.pack(fill="x", pady=(0, 10))
            ttk.Label(
                tip_panel,
                text="分组规则按最终输出款号匹配。特殊规则用于把指定款号直接归到某个分组，例如 SK-010 -> MG；多个款号可用逗号分隔。",
                style="Muted.TLabel",
                wraplength=1050,
                justify="left",
            ).pack(fill="x")

            cols = ("enabled", "priority", "name", "match_type", "pattern", "group")
            table = self._panel(container)
            table.pack(fill="both", expand=True)
            self.group_tree = ttk.Treeview(table, columns=cols, show="headings")
            headings = {
                "enabled": "启用", "priority": "优先级", "name": "规则名", "match_type": "匹配方式",
                "pattern": "匹配款号", "group": "分组名称"
            }
            widths = {"enabled": 60, "priority": 80, "name": 260, "match_type": 120, "pattern": 260, "group": 180}
            for col in cols:
                self.group_tree.heading(col, text=headings[col])
                self.group_tree.column(col, width=widths[col], anchor="center", stretch=False)
            self.group_tree.grid(row=0, column=0, sticky="nsew")
            group_sb = ttk.Scrollbar(table, orient="vertical", command=self.group_tree.yview)
            group_xsb = ttk.Scrollbar(table, orient="horizontal", command=self.group_tree.xview)
            self.group_tree.configure(yscrollcommand=group_sb.set, xscrollcommand=group_xsb.set)
            group_sb.grid(row=0, column=1, sticky="ns")
            group_xsb.grid(row=1, column=0, sticky="ew")
            table.rowconfigure(0, weight=1)
            table.columnconfigure(0, weight=1)
            self.group_tree.bind("<Double-1>", lambda e: self.edit_group_rule())
            self.group_tree.bind("<Delete>", lambda e: self.delete_group_rule())

        def _build_color_tab(self):
            main = ttk.Frame(self.tab_color, style="App.TFrame", padding=(0, 12, 0, 0))
            main.pack(fill="both", expand=True)
            for r in range(2):
                main.rowconfigure(r, weight=1)
            for c in range(2):
                main.columnconfigure(c, weight=1)

            color_panel = self._panel(main)
            color_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 6), pady=(0, 6))
            color_header = ttk.Frame(color_panel, style="Panel.TFrame")
            color_header.pack(fill="x", pady=(0, 8))
            ttk.Label(color_header, text="英文颜色 → 中文颜色", style="PanelTitle.TLabel").pack(side="left")
            ttk.Button(color_header, text="新增/修改", command=self.edit_color).pack(side="right", padx=(6, 0))
            ttk.Button(color_header, text="删除", command=self.delete_color, style="Danger.TButton").pack(side="right")
            self.color_tree = ttk.Treeview(color_panel, columns=("en", "cn"), show="headings")
            self.color_tree.heading("en", text="英文颜色")
            self.color_tree.heading("cn", text="中文颜色")
            self.color_tree.column("en", width=150, anchor="center")
            self.color_tree.column("cn", width=150, anchor="center")
            self.color_tree.pack(side="left", fill="both", expand=True)
            color_sb = ttk.Scrollbar(color_panel, orient="vertical", command=self.color_tree.yview)
            self.color_tree.configure(yscrollcommand=color_sb.set)
            color_sb.pack(side="right", fill="y")
            self.color_tree.bind("<Double-1>", lambda e: self.edit_color())
            self.color_tree.bind("<Delete>", lambda e: self.delete_color())

            prefix_panel = self._panel(main)
            prefix_panel.grid(row=0, column=1, sticky="nsew", padx=(6, 0), pady=(0, 6))
            prefix_header = ttk.Frame(prefix_panel, style="Panel.TFrame")
            prefix_header.pack(fill="x", pady=(0, 8))
            ttk.Label(prefix_header, text="款号前缀替换", style="PanelTitle.TLabel").pack(side="left")
            ttk.Button(prefix_header, text="新增/修改", command=self.edit_prefix).pack(side="right", padx=(6, 0))
            ttk.Button(prefix_header, text="删除", command=self.delete_prefix, style="Danger.TButton").pack(side="right")
            self.prefix_tree = ttk.Treeview(prefix_panel, columns=("enabled", "from", "to"), show="headings")
            self.prefix_tree.heading("enabled", text="启用")
            self.prefix_tree.heading("from", text="原前缀")
            self.prefix_tree.heading("to", text="替换为")
            self.prefix_tree.column("enabled", width=60, anchor="center")
            self.prefix_tree.column("from", width=140, anchor="center")
            self.prefix_tree.column("to", width=140, anchor="center")
            self.prefix_tree.pack(side="left", fill="both", expand=True)
            prefix_sb = ttk.Scrollbar(prefix_panel, orient="vertical", command=self.prefix_tree.yview)
            self.prefix_tree.configure(yscrollcommand=prefix_sb.set)
            prefix_sb.pack(side="right", fill="y")
            self.prefix_tree.bind("<Double-1>", lambda e: self.edit_prefix())
            self.prefix_tree.bind("<Delete>", lambda e: self.delete_prefix())

            style_panel = self._panel(main)
            style_panel.grid(row=1, column=0, sticky="nsew", padx=(0, 6), pady=(6, 0))
            style_header = ttk.Frame(style_panel, style="Panel.TFrame")
            style_header.pack(fill="x", pady=(0, 8))
            ttk.Label(style_header, text="款号对照替换", style="PanelTitle.TLabel").pack(side="left")
            ttk.Button(style_header, text="新增/修改", command=self.edit_style_code_replace).pack(side="right", padx=(6, 0))
            ttk.Button(style_header, text="删除", command=self.delete_style_code_replace, style="Danger.TButton").pack(side="right")
            self.style_code_tree = ttk.Treeview(style_panel, columns=("enabled", "from", "to"), show="headings")
            self.style_code_tree.heading("enabled", text="启用")
            self.style_code_tree.heading("from", text="原款号")
            self.style_code_tree.heading("to", text="替换为")
            self.style_code_tree.column("enabled", width=60, anchor="center")
            self.style_code_tree.column("from", width=140, anchor="center")
            self.style_code_tree.column("to", width=140, anchor="center")
            self.style_code_tree.pack(side="left", fill="both", expand=True)
            style_sb = ttk.Scrollbar(style_panel, orient="vertical", command=self.style_code_tree.yview)
            self.style_code_tree.configure(yscrollcommand=style_sb.set)
            style_sb.pack(side="right", fill="y")
            self.style_code_tree.bind("<Double-1>", lambda e: self.edit_style_code_replace())
            self.style_code_tree.bind("<Delete>", lambda e: self.delete_style_code_replace())

            quantity_panel = self._panel(main)
            quantity_panel.grid(row=1, column=1, sticky="nsew", padx=(6, 0), pady=(6, 0))
            quantity_header = ttk.Frame(quantity_panel, style="Panel.TFrame")
            quantity_header.pack(fill="x", pady=(0, 8))
            ttk.Label(quantity_header, text="置顶数量款号", style="PanelTitle.TLabel").pack(side="left")
            ttk.Button(quantity_header, text="保存配置", command=self.save_all_config, style="Accent.TButton").pack(side="right", padx=(6, 0))
            ttk.Button(quantity_header, text="新增/修改", command=self.edit_quantity_only_style).pack(side="right", padx=(6, 0))
            ttk.Button(quantity_header, text="删除", command=self.delete_quantity_only_style, style="Danger.TButton").pack(side="right")
            ttk.Label(
                quantity_panel,
                text="这些最终款号会放在汇总表最前；置顶分组标题只显示数量，底部总款数仍会计入。",
                style="Muted.TLabel",
                wraplength=420,
                justify="left",
            ).pack(fill="x", pady=(0, 6))
            self.quantity_only_tree = ttk.Treeview(quantity_panel, columns=("style",), show="headings")
            self.quantity_only_tree.heading("style", text="最终款号")
            self.quantity_only_tree.column("style", width=220, anchor="center")
            self.quantity_only_tree.pack(side="left", fill="both", expand=True)
            quantity_sb = ttk.Scrollbar(quantity_panel, orient="vertical", command=self.quantity_only_tree.yview)
            self.quantity_only_tree.configure(yscrollcommand=quantity_sb.set)
            quantity_sb.pack(side="right", fill="y")
            self.quantity_only_tree.bind("<Double-1>", lambda e: self.edit_quantity_only_style())
            self.quantity_only_tree.bind("<Delete>", lambda e: self.delete_quantity_only_style())

        def add_log(self, text: str):
            self.log.insert("end", text + "\n")
            self.log.see("end")
            self.update_idletasks()

        def clear_log(self):
            self.log.delete("1.0", "end")
            self._set_status("日志已清空")

        def add_pdfs(self):
            paths = filedialog.askopenfilenames(title="选择PDF", filetypes=[("PDF文件", "*.pdf")])
            self._add_pdf_paths([Path(p) for p in paths])

        def add_folder(self):
            folder = filedialog.askdirectory(title="选择PDF文件夹")
            if folder:
                self._add_pdf_paths(collect_pdf_paths([Path(folder)]))

        def _add_pdf_paths(self, paths: Iterable[Path]):
            was_empty = not self.pdf_paths
            existing = {str(p.resolve()) for p in self.pdf_paths}
            added = 0
            for p in paths:
                if p.exists() and p.suffix.lower() == ".pdf" and str(p.resolve()) not in existing:
                    self.pdf_paths.append(p)
                    existing.add(str(p.resolve()))
                    added += 1
            self.refresh_pdf_tree()
            if added:
                if was_empty:
                    self.output_var.set(str(self._default_output_path()))
                self._set_status(f"已添加 {added} 个 PDF")
            else:
                self._set_status("没有新增 PDF")

        def refresh_pdf_tree(self):
            for i in self.pdf_tree.get_children():
                self.pdf_tree.delete(i)
            for idx, p in enumerate(self.pdf_paths):
                self.pdf_tree.insert("", "end", iid=f"pdf:{idx}", values=(p.name, str(p.parent), str(p)))
            self.pdf_count_var.set(f"{len(self.pdf_paths)} 个 PDF")
            state = "normal" if self.pdf_paths and not self.processing else "disabled"
            if hasattr(self, "run_btn"):
                self.run_btn.configure(state=state)
            if hasattr(self, "remove_pdf_btn"):
                self.remove_pdf_btn.configure(state="normal" if self.pdf_paths else "disabled")
            if hasattr(self, "clear_pdf_btn"):
                self.clear_pdf_btn.configure(state="normal" if self.pdf_paths else "disabled")

        def remove_selected_pdf(self):
            selected = self.pdf_tree.selection()
            if not selected:
                self._set_status("请先选择要移除的 PDF")
                return
            selected_paths = {self.pdf_tree.item(i, "values")[2] for i in selected}
            self.pdf_paths = [p for p in self.pdf_paths if str(p) not in selected_paths]
            self.refresh_pdf_tree()
            self._set_status(f"已移除 {len(selected)} 个 PDF")

        def clear_pdfs(self):
            self.pdf_paths = []
            self.refresh_pdf_tree()
            self._set_status("PDF 队列已清空")

        def choose_output(self):
            path = filedialog.asksaveasfilename(
                title="保存输出Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                initialdir=str((self.pdf_paths[0].parent if self.pdf_paths else Path.cwd()).resolve()),
                initialfile=Path(self.output_var.get() or self._default_output_path()).name,
            )
            if path:
                self.output_var.set(path)
                self._set_status("已更新输出位置")

        def open_output_folder(self):
            path = self.last_output_path or Path(self.output_var.get().strip())
            folder = path.parent if path.suffix else path
            if not folder.exists():
                messagebox.showwarning("提示", "输出目录不存在。")
                return
            try:
                os.startfile(str(folder))  # type: ignore[attr-defined]
            except Exception as e:
                messagebox.showerror("打开失败", str(e))

        def run_process(self):
            if not self.pdf_paths:
                messagebox.showwarning("提示", "请先添加至少一个 PDF。")
                return
            out_path = Path(self.output_var.get().strip())
            if not str(out_path):
                messagebox.showwarning("提示", "请设置输出 Excel 路径。")
                return
            if out_path.suffix.lower() != ".xlsx":
                messagebox.showwarning("提示", "输出路径必须是 .xlsx 文件。")
                return
            try:
                self._set_busy(True)
                self._set_status("正在处理 PDF...")
                self.save_all_config(show_msg=False)
                out_path.parent.mkdir(parents=True, exist_ok=True)
                self.add_log("开始处理...")
                result = process_pdfs(self.pdf_paths, out_path, self.cfg)
                self.last_output_path = out_path
                self.add_log(f"完成：{result['out_path']}")
                self.add_log(f"PDF数量：{result['pdf_count']}")
                self.add_log(f"解析行数：{result['parsed_rows']}")
                self.add_log(f"PDF原始数量合计：{result['original_qty']}")
                self.add_log(f"输出汇总款数：{result['output_lines']}款")
                self.add_log(f"输出汇总数量：{result['output_qty']}")
                self.add_log(f"异常数量：{result['issues']}")
                self.result_summary_var.set(
                    f"上次处理：{result['pdf_count']} 个 PDF，解析 {result['parsed_rows']} 行，原始 {result['original_qty']}，输出 {result['output_qty']}，异常 {result['issues']}"
                )
                self.config_state_var.set(f"配置：{self.cfg_path.name}")
                self.open_output_btn.configure(state="normal")
                self._set_status("处理完成")
                messagebox.showinfo("完成", f"已生成：\n{result['out_path']}")
            except Exception as e:
                self.add_log("处理失败：" + str(e))
                self.add_log(traceback.format_exc())
                self._set_status("处理失败")
                messagebox.showerror("处理失败", str(e))
            finally:
                self._set_busy(False)
                self.refresh_pdf_tree()

        def refresh_rules_tree(self):
            selected_idx = self._selected_rule_actual_index() if self.rules_tree.selection() else None
            for i in self.rules_tree.get_children():
                self.rules_tree.delete(i)
            for idx, rule in sorted_indexed_rules(self.cfg):
                self.rules_tree.insert("", "end", iid=f"rule:{idx}", values=(
                    "启用" if rule.get("enabled", True) else "停用",
                    rule.get("priority", ""),
                    rule.get("name", ""),
                ))
            if selected_idx is not None and f"rule:{selected_idx}" in self.rules_tree.get_children():
                self.rules_tree.selection_set(f"rule:{selected_idx}")
            elif self.rules_tree.get_children():
                first = self.rules_tree.get_children()[0]
                self.rules_tree.selection_set(first)
            self.refresh_rule_detail()

        def refresh_rule_detail(self):
            idx = self._selected_rule_actual_index()
            if idx is None:
                self.rule_detail_title_var.set("请选择一条规则")
                self.rule_detail_meta_var.set("")
                for var in self.rule_detail_fields.values():
                    var.set("")
                return

            rule = self.cfg.get("rules", [])[idx]
            enabled = "启用" if rule.get("enabled", True) else "停用"
            self.rule_detail_title_var.set(str(rule.get("name", "未命名规则")))
            self.rule_detail_meta_var.set(
                f"{enabled}  |  优先级 {rule.get('priority', '')}  |  命中后输出 {rule.get('output_styles', '{auto}') or '{auto}'}"
            )
            for key, var in self.rule_detail_fields.items():
                value = rule.get(key, "")
                if key == "fixed_colors" and not value:
                    value = "无"
                if key == "group_name" and not value:
                    value = "按分组规则 / 未分组"
                if key == "note" and not value:
                    value = "无"
                var.set(str(value))

        def _selected_rule_actual_index(self) -> Optional[int]:
            sel = self.rules_tree.selection()
            if not sel:
                return None
            try:
                idx = int(str(sel[0]).split(":", 1)[1])
            except (IndexError, ValueError):
                return None
            return idx if 0 <= idx < len(self.cfg.get("rules", [])) else None

        def add_rule(self):
            dlg = RuleDialog(self)
            self.wait_window(dlg)
            if dlg.result:
                self.cfg.setdefault("rules", []).append(dlg.result)
                self.refresh_rules_tree()
                new_iid = f"rule:{len(self.cfg.get('rules', [])) - 1}"
                if new_iid in self.rules_tree.get_children():
                    self.rules_tree.selection_set(new_iid)
                    self.rules_tree.see(new_iid)
                    self.refresh_rule_detail()
                self._mark_config_dirty()
                self._set_status("已新增规则")

        def edit_rule(self):
            idx = self._selected_rule_actual_index()
            if idx is None:
                messagebox.showwarning("提示", "请先选择一条规则。")
                return
            dlg = RuleDialog(self, dict(self.cfg["rules"][idx]))
            self.wait_window(dlg)
            if dlg.result:
                self.cfg["rules"][idx] = dlg.result
                self.refresh_rules_tree()
                iid = f"rule:{idx}"
                if iid in self.rules_tree.get_children():
                    self.rules_tree.selection_set(iid)
                    self.rules_tree.see(iid)
                    self.refresh_rule_detail()
                self._mark_config_dirty()
                self._set_status("已更新规则")

        def delete_rule(self):
            idx = self._selected_rule_actual_index()
            if idx is None:
                self._set_status("请先选择要删除的规则")
                return
            if messagebox.askyesno("确认", "确定删除选中的规则吗？"):
                self.cfg["rules"].pop(idx)
                self.refresh_rules_tree()
                self._mark_config_dirty()
                self._set_status("已删除规则")

        def refresh_group_rules_tree(self):
            for i in self.group_tree.get_children():
                self.group_tree.delete(i)
            for idx, rule in sorted_indexed_group_rules(self.cfg):
                self.group_tree.insert("", "end", iid=f"group:{idx}", values=(
                    "是" if rule.get("enabled", True) else "否",
                    rule.get("priority", ""),
                    rule.get("name", ""),
                    rule.get("match_type", ""),
                    rule.get("pattern", ""),
                    rule.get("group_name", ""),
                ))

        def _selected_group_rule_actual_index(self) -> Optional[int]:
            sel = self.group_tree.selection()
            if not sel:
                return None
            try:
                idx = int(str(sel[0]).split(":", 1)[1])
            except (IndexError, ValueError):
                return None
            return idx if 0 <= idx < len(self.cfg.get("group_rules", [])) else None

        def add_group_rule(self):
            dlg = GroupRuleDialog(self)
            self.wait_window(dlg)
            if dlg.result:
                self.cfg.setdefault("group_rules", []).append(dlg.result)
                self.refresh_group_rules_tree()
                self._mark_config_dirty()
                self._set_status("已新增分组规则")

        def add_special_group_rule(self):
            dlg = GroupRuleDialog(self, {
                "enabled": True,
                "priority": 10,
                "name": "特殊分组规则",
                "match_type": "exact",
                "pattern": "",
                "group_name": "",
            })
            self.wait_window(dlg)
            if dlg.result:
                self.cfg.setdefault("group_rules", []).append(dlg.result)
                self.refresh_group_rules_tree()
                self._mark_config_dirty()
                self._set_status("已新增特殊分组规则")

        def edit_group_rule(self):
            idx = self._selected_group_rule_actual_index()
            if idx is None:
                messagebox.showwarning("提示", "请先选择一条分组规则。")
                return
            dlg = GroupRuleDialog(self, dict(self.cfg["group_rules"][idx]))
            self.wait_window(dlg)
            if dlg.result:
                self.cfg["group_rules"][idx] = dlg.result
                self.refresh_group_rules_tree()
                self._mark_config_dirty()
                self._set_status("已更新分组规则")

        def delete_group_rule(self):
            idx = self._selected_group_rule_actual_index()
            if idx is None:
                self._set_status("请先选择要删除的分组规则")
                return
            if messagebox.askyesno("确认", "确定删除选中的分组规则吗？"):
                self.cfg["group_rules"].pop(idx)
                self.refresh_group_rules_tree()
                self._mark_config_dirty()
                self._set_status("已删除分组规则")

        def refresh_color_table(self):
            for i in self.color_tree.get_children():
                self.color_tree.delete(i)
            for en, cn in sorted(self.cfg.get("color_map", {}).items()):
                self.color_tree.insert("", "end", values=(en, cn))

        def edit_color(self):
            sel = self.color_tree.selection()
            old_en = old_cn = ""
            if sel:
                vals = self.color_tree.item(sel[0], "values")
                old_en, old_cn = vals[0], vals[1]
            win = tk.Toplevel(self)
            win.title("颜色映射")
            win.resizable(False, False)
            en_var = tk.StringVar(value=old_en)
            cn_var = tk.StringVar(value=old_cn)
            ttk.Label(win, text="英文颜色").grid(row=0, column=0, padx=8, pady=6, sticky="e")
            ttk.Entry(win, textvariable=en_var, width=28).grid(row=0, column=1, padx=8, pady=6)
            ttk.Label(win, text="中文颜色").grid(row=1, column=0, padx=8, pady=6, sticky="e")
            ttk.Entry(win, textvariable=cn_var, width=28).grid(row=1, column=1, padx=8, pady=6)
            def ok():
                en = en_var.get().strip()
                cn = cn_var.get().strip()
                if not en or not cn:
                    messagebox.showerror("错误", "英文颜色和中文颜色不能为空。")
                    return
                if old_en and old_en != en:
                    self.cfg["color_map"].pop(old_en, None)
                self.cfg.setdefault("color_map", {})[en] = cn
                self.refresh_color_table()
                self._mark_config_dirty()
                self._set_status("已更新颜色映射")
                win.destroy()
            ttk.Button(win, text="保存", command=ok).grid(row=2, column=0, columnspan=2, pady=8)
            win.grab_set()

        def delete_color(self):
            sel = self.color_tree.selection()
            if not sel:
                self._set_status("请先选择要删除的颜色")
                return
            en = self.color_tree.item(sel[0], "values")[0]
            self.cfg.get("color_map", {}).pop(en, None)
            self.refresh_color_table()
            self._mark_config_dirty()
            self._set_status("已删除颜色映射")

        def refresh_prefix_table(self):
            for i in self.prefix_tree.get_children():
                self.prefix_tree.delete(i)
            for item in self.cfg.get("general", {}).get("style_prefix_replace", []):
                self.prefix_tree.insert("", "end", values=("是" if item.get("enabled", True) else "否", item.get("from", ""), item.get("to", "")))

        def edit_prefix(self):
            sel = self.prefix_tree.selection()
            actual_idx = None
            enabled = True
            old_from = old_to = ""
            if sel:
                vals = self.prefix_tree.item(sel[0], "values")
                enabled = vals[0] == "是"
                old_from, old_to = vals[1], vals[2]
                for idx, item in enumerate(self.cfg.get("general", {}).get("style_prefix_replace", [])):
                    if item.get("from", "") == old_from and item.get("to", "") == old_to:
                        actual_idx = idx
                        break
            win = tk.Toplevel(self)
            win.title("款号前缀替换")
            win.resizable(False, False)
            enabled_var = tk.BooleanVar(value=enabled)
            from_var = tk.StringVar(value=old_from)
            to_var = tk.StringVar(value=old_to)
            ttk.Checkbutton(win, text="启用", variable=enabled_var).grid(row=0, column=0, columnspan=2, padx=8, pady=6, sticky="w")
            ttk.Label(win, text="原前缀").grid(row=1, column=0, padx=8, pady=6, sticky="e")
            ttk.Entry(win, textvariable=from_var, width=28).grid(row=1, column=1, padx=8, pady=6)
            ttk.Label(win, text="替换为").grid(row=2, column=0, padx=8, pady=6, sticky="e")
            ttk.Entry(win, textvariable=to_var, width=28).grid(row=2, column=1, padx=8, pady=6)
            def ok():
                src = from_var.get().strip()
                dst = to_var.get().strip()
                if not src:
                    messagebox.showerror("错误", "原前缀不能为空。")
                    return
                item = {"enabled": bool(enabled_var.get()), "from": src, "to": dst}
                lst = self.cfg.setdefault("general", {}).setdefault("style_prefix_replace", [])
                if actual_idx is None:
                    lst.append(item)
                else:
                    lst[actual_idx] = item
                self.refresh_prefix_table()
                self._mark_config_dirty()
                self._set_status("已更新款号前缀替换")
                win.destroy()
            ttk.Button(win, text="保存", command=ok).grid(row=3, column=0, columnspan=2, pady=8)
            win.grab_set()

        def delete_prefix(self):
            sel = self.prefix_tree.selection()
            if not sel:
                self._set_status("请先选择要删除的前缀替换")
                return
            vals = self.prefix_tree.item(sel[0], "values")
            old_from, old_to = vals[1], vals[2]
            lst = self.cfg.get("general", {}).get("style_prefix_replace", [])
            self.cfg["general"]["style_prefix_replace"] = [x for x in lst if not (x.get("from") == old_from and x.get("to") == old_to)]
            self.refresh_prefix_table()
            self._mark_config_dirty()
            self._set_status("已删除款号前缀替换")

        def refresh_style_code_table(self):
            for i in self.style_code_tree.get_children():
                self.style_code_tree.delete(i)
            for item in self.cfg.get("general", {}).get("style_code_replace", []):
                self.style_code_tree.insert("", "end", values=("是" if item.get("enabled", True) else "否", item.get("from", ""), item.get("to", "")))

        def edit_style_code_replace(self):
            sel = self.style_code_tree.selection()
            actual_idx = None
            enabled = True
            old_from = old_to = ""
            if sel:
                vals = self.style_code_tree.item(sel[0], "values")
                enabled = vals[0] == "是"
                old_from, old_to = vals[1], vals[2]
                for idx, item in enumerate(self.cfg.get("general", {}).get("style_code_replace", [])):
                    if item.get("from", "") == old_from and item.get("to", "") == old_to:
                        actual_idx = idx
                        break
            win = tk.Toplevel(self)
            win.title("款号对照替换")
            win.resizable(False, False)
            enabled_var = tk.BooleanVar(value=enabled)
            from_var = tk.StringVar(value=old_from)
            to_var = tk.StringVar(value=old_to)
            ttk.Checkbutton(win, text="启用", variable=enabled_var).grid(row=0, column=0, columnspan=2, padx=8, pady=6, sticky="w")
            ttk.Label(win, text="原款号").grid(row=1, column=0, padx=8, pady=6, sticky="e")
            ttk.Entry(win, textvariable=from_var, width=28).grid(row=1, column=1, padx=8, pady=6)
            ttk.Label(win, text="替换为").grid(row=2, column=0, padx=8, pady=6, sticky="e")
            ttk.Entry(win, textvariable=to_var, width=28).grid(row=2, column=1, padx=8, pady=6)

            def ok():
                src = from_var.get().strip()
                dst = to_var.get().strip()
                if not src or not dst:
                    messagebox.showerror("错误", "原款号和替换款号不能为空。")
                    return
                item = {"enabled": bool(enabled_var.get()), "from": src, "to": dst}
                lst = self.cfg.setdefault("general", {}).setdefault("style_code_replace", [])
                if actual_idx is None:
                    lst.append(item)
                else:
                    lst[actual_idx] = item
                self.refresh_style_code_table()
                self._mark_config_dirty()
                self._set_status("已更新款号对照替换")
                win.destroy()

            ttk.Button(win, text="保存", command=ok).grid(row=3, column=0, columnspan=2, pady=8)
            win.grab_set()

        def delete_style_code_replace(self):
            sel = self.style_code_tree.selection()
            if not sel:
                self._set_status("请先选择要删除的款号对照")
                return
            vals = self.style_code_tree.item(sel[0], "values")
            old_from, old_to = vals[1], vals[2]
            lst = self.cfg.get("general", {}).get("style_code_replace", [])
            self.cfg["general"]["style_code_replace"] = [x for x in lst if not (x.get("from") == old_from and x.get("to") == old_to)]
            self.refresh_style_code_table()
            self._mark_config_dirty()
            self._set_status("已删除款号对照替换")

        def refresh_quantity_only_table(self):
            for i in self.quantity_only_tree.get_children():
                self.quantity_only_tree.delete(i)
            styles = self.cfg.get("general", {}).get("quantity_only_styles", [])
            if isinstance(styles, str):
                styles = normalize_separator_list(styles)
            for style in styles:
                self.quantity_only_tree.insert("", "end", values=(style,))

        def edit_quantity_only_style(self):
            sel = self.quantity_only_tree.selection()
            old_style = ""
            if sel:
                old_style = self.quantity_only_tree.item(sel[0], "values")[0]
            win = tk.Toplevel(self)
            win.title("置顶数量款号")
            win.resizable(False, False)
            style_var = tk.StringVar(value=old_style)
            ttk.Label(win, text="最终款号").grid(row=0, column=0, padx=8, pady=6, sticky="e")
            ttk.Entry(win, textvariable=style_var, width=28).grid(row=0, column=1, padx=8, pady=6)

            def ok():
                style = style_var.get().strip()
                if not style:
                    messagebox.showerror("错误", "款号不能为空。")
                    return
                styles = self.cfg.setdefault("general", {}).setdefault("quantity_only_styles", [])
                if isinstance(styles, str):
                    styles = normalize_separator_list(styles)
                    self.cfg["general"]["quantity_only_styles"] = styles
                if old_style and old_style in styles:
                    styles[styles.index(old_style)] = style
                elif style not in styles:
                    styles.append(style)
                self.refresh_quantity_only_table()
                self._mark_config_dirty()
                self._set_status("已更新置顶数量款号")
                win.destroy()

            ttk.Button(win, text="保存", command=ok).grid(row=1, column=0, columnspan=2, pady=8)
            win.grab_set()

        def delete_quantity_only_style(self):
            sel = self.quantity_only_tree.selection()
            if not sel:
                self._set_status("请先选择要删除的置顶数量款号")
                return
            style = self.quantity_only_tree.item(sel[0], "values")[0]
            styles = self.cfg.get("general", {}).get("quantity_only_styles", [])
            if isinstance(styles, str):
                styles = normalize_separator_list(styles)
            self.cfg["general"]["quantity_only_styles"] = [x for x in styles if x != style]
            self.refresh_quantity_only_table()
            self._mark_config_dirty()
            self._set_status("已删除置顶数量款号")

        def save_all_config(self, show_msg: bool = True):
            save_config(self.cfg, self.cfg_path)
            self.config_state_var.set(f"配置：{self.cfg_path.name}")
            self._set_status("配置已保存")
            if show_msg:
                messagebox.showinfo("已保存", f"配置已保存到：\n{self.cfg_path}")

        def import_config(self):
            path = filedialog.askopenfilename(title="导入配置", filetypes=[("JSON配置", "*.json")])
            if path:
                self.cfg = load_config(Path(path))
                self.refresh_rules_tree()
                self.refresh_group_rules_tree()
                self.refresh_color_table()
                self.refresh_prefix_table()
                self.refresh_style_code_table()
                self.refresh_quantity_only_table()
                self._mark_config_dirty()
                self._set_status("配置已导入，尚未保存到默认配置文件")
                messagebox.showinfo("完成", "配置已导入。点击“保存配置”后会写入当前默认配置文件。")

        def export_config(self):
            path = filedialog.asksaveasfilename(title="导出配置", defaultextension=".json", filetypes=[("JSON配置", "*.json")])
            if path:
                save_config(self.cfg, Path(path))
                self._set_status("配置已导出")
                messagebox.showinfo("完成", f"配置已导出：\n{path}")

        def reset_config(self):
            if messagebox.askyesno("确认", "确定恢复默认配置吗？当前未保存的规则会被覆盖。"):
                self.cfg = deep_copy_default_config()
                self.refresh_rules_tree()
                self.refresh_group_rules_tree()
                self.refresh_color_table()
                self.refresh_prefix_table()
                self.refresh_style_code_table()
                self.refresh_quantity_only_table()
                self.output_var.set(str(self._default_output_path()))
                self._mark_config_dirty()
                self._set_status("已恢复默认配置，尚未保存")


def run_cli(args: argparse.Namespace) -> None:
    cfg_path = Path(args.config) if args.config else default_config_path()
    cfg = load_config(cfg_path)
    inputs = [Path(x) for x in args.inputs]
    pdfs = collect_pdf_paths(inputs)
    if not pdfs:
        raise SystemExit("没有找到 PDF 文件。")

    out_arg = Path(args.output) if args.output else None
    if out_arg is None:
        out_path = pdfs[0].with_name(pdfs[0].stem + "_报货汇总.xlsx") if len(pdfs) == 1 else Path.cwd() / "报货汇总.xlsx"
    elif out_arg.suffix.lower() == ".xlsx":
        out_path = out_arg
    else:
        out_arg.mkdir(parents=True, exist_ok=True)
        out_path = out_arg / (pdfs[0].stem + "_报货汇总.xlsx" if len(pdfs) == 1 else "报货汇总.xlsx")

    result = process_pdfs(pdfs, out_path, cfg)
    print("完成：", result["out_path"])
    print("PDF数量：", result["pdf_count"])
    print("解析行数：", result["parsed_rows"])
    print("PDF原始数量合计：", result["original_qty"])
    print("输出汇总款数：", result["output_lines"], "款")
    print("输出汇总数量：", result["output_qty"])
    print("异常数量：", result["issues"])


def main():
    parser = argparse.ArgumentParser(description="报货 PDF 自动汇总工具")
    parser.add_argument("--cli", action="store_true", help="使用命令行模式，不打开 GUI")
    parser.add_argument("inputs", nargs="*", help="PDF 文件或包含 PDF 的文件夹")
    parser.add_argument("-o", "--output", help="输出 Excel 文件或输出目录")
    parser.add_argument("--config", help="配置 JSON 路径")
    args = parser.parse_args()

    if args.cli:
        run_cli(args)
        return

    if tk is None:
        raise SystemExit("当前 Python 环境没有 tkinter，无法启动 GUI。可以使用 --cli 命令行模式。")
    enable_dpi_awareness()
    app = BaohuoApp()
    app.mainloop()


if __name__ == "__main__":
    main()
